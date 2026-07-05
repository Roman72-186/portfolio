"""Tests for /cabinet/superadmin — dashboard, set-credentials, issue-link."""
from datetime import datetime, timezone

import pytest

from app.models.feedback import Feedback, FeedbackMessage
from app.models.login_token import LoginToken
from app.models.mock_exam_attempt import MockExamAttempt
from app.models.user import User
from app.models.work import Work, WORK_TYPE_MOCK_EXAM
from app.services.feedback import ROLE_CURATOR


# ---------------------------------------------------------------------------
# Fixtures
# ---------------------------------------------------------------------------

@pytest.fixture()
def superadmin_client(client, db, user_factory, session_factory):
    user = user_factory(vk_id=900001, name="Super Admin", role_name="суперадмин")
    sess = session_factory(user)
    client.cookies.set("session_id", sess.id)
    return client, user


@pytest.fixture()
def curator_user(db, user_factory):
    user = user_factory(vk_id=900002, name="Curator User", role_name="куратор")
    user.first_name = "Иван"
    user.last_name = "Петров"
    db.add(user)
    db.commit()
    db.refresh(user)
    return user


# ---------------------------------------------------------------------------
# GET /cabinet/superadmin
# ---------------------------------------------------------------------------

def test_superadmin_dashboard_loads(superadmin_client):
    client, _ = superadmin_client
    resp = client.get("/cabinet/superadmin")
    assert resp.status_code == 200


def test_superadmin_dashboard_contains_stats(superadmin_client):
    client, _ = superadmin_client
    resp = client.get("/cabinet/superadmin")
    assert resp.status_code == 200
    # Should contain some dashboard content
    text = resp.text.lower()
    assert "пользовател" in text or "куратор" in text or "суперадмин" in text


def test_superadmin_dashboard_denied_for_curator(client, db, user_factory, session_factory):
    user = user_factory(vk_id=900010, name="Just Curator", role_name="куратор")
    sess = session_factory(user)
    client.cookies.set("session_id", sess.id)
    resp = client.get("/cabinet/superadmin", follow_redirects=False)
    assert resp.status_code == 403


def test_superadmin_dashboard_denied_for_student(client, db, user_factory, session_factory):
    user = user_factory(vk_id=900011, name="Student", role_name="ученик")
    sess = session_factory(user)
    client.cookies.set("session_id", sess.id)
    resp = client.get("/cabinet/superadmin", follow_redirects=False)
    assert resp.status_code == 403


def test_superadmin_dashboard_denied_no_session(client):
    resp = client.get("/cabinet/superadmin", follow_redirects=False)
    assert resp.status_code in (302, 401)


# ---------------------------------------------------------------------------
# POST /cabinet/superadmin/set-credentials
# ---------------------------------------------------------------------------

def test_set_credentials_generates_login(superadmin_client, db, curator_user):
    client, _ = superadmin_client
    resp = client.post("/cabinet/superadmin/set-credentials",
                       data={"target_user_id": curator_user.id, "csrf_token": "bypass"})
    assert resp.status_code == 200

    db.refresh(curator_user)
    assert curator_user.staff_login is not None
    assert curator_user.password_hash is not None


def test_set_credentials_shows_issued_creds(superadmin_client, db, curator_user):
    client, _ = superadmin_client
    resp = client.post("/cabinet/superadmin/set-credentials",
                       data={"target_user_id": curator_user.id, "csrf_token": "bypass"})
    assert resp.status_code == 200
    text = resp.text.lower()
    assert "логин" in text or "пароль" in text or "login" in text


def test_set_credentials_reuses_existing_login(superadmin_client, db, curator_user):
    """If staff_login already set, keep it; only reset password."""
    curator_user.staff_login = "existing.login"
    db.add(curator_user)
    db.commit()

    client, _ = superadmin_client
    client.post("/cabinet/superadmin/set-credentials",
                data={"target_user_id": curator_user.id, "csrf_token": "bypass"})

    db.refresh(curator_user)
    assert curator_user.staff_login == "existing.login"


def test_set_credentials_nonexistent_user_404(superadmin_client):
    client, _ = superadmin_client
    resp = client.post("/cabinet/superadmin/set-credentials",
                       data={"target_user_id": 99999, "csrf_token": "bypass"})
    assert resp.status_code == 404


def test_set_credentials_denied_for_curator(client, db, user_factory, session_factory, curator_user):
    curator = user_factory(vk_id=900020, name="Curator", role_name="куратор")
    sess = session_factory(curator)
    client.cookies.set("session_id", sess.id)
    resp = client.post("/cabinet/superadmin/set-credentials",
                       data={"target_user_id": curator_user.id, "csrf_token": "bypass"},
                       follow_redirects=False)
    assert resp.status_code == 403


def test_set_credentials_succeeds_for_admin(client, db, user_factory, session_factory, curator_user):
    admin = user_factory(vk_id=900021, name="Admin Actor", role_name="админ", is_admin=True)
    sess = session_factory(admin)
    client.cookies.set("session_id", sess.id)
    resp = client.post("/cabinet/superadmin/set-credentials",
                       data={"target_user_id": curator_user.id, "csrf_token": "bypass"},
                       follow_redirects=False)
    assert resp.status_code == 200

    db.refresh(curator_user)
    assert curator_user.staff_login is not None
    assert curator_user.password_hash is not None


def test_set_credentials_denies_peer_rank_for_admin(client, db, user_factory, session_factory):
    admin = user_factory(vk_id=900022, name="Admin Actor", role_name="админ", is_admin=True)
    peer = user_factory(vk_id=900023, name="Admin Peer", role_name="админ")
    sess = session_factory(admin)
    client.cookies.set("session_id", sess.id)
    resp = client.post("/cabinet/superadmin/set-credentials",
                       data={"target_user_id": peer.id, "csrf_token": "bypass"},
                       follow_redirects=False)
    assert resp.status_code == 403

    db.refresh(peer)
    assert peer.staff_login is None
    assert peer.password_hash is None


def test_superadmin_create_manual_student_issues_credentials(superadmin_client, db, curator_user):
    client, _ = superadmin_client

    resp = client.post(
        "/cabinet/superadmin/users/create-student",
        data={
            "first_name": "Анна",
            "last_name": "Иванова",
            "tg_username": "@anna_arch",
            "tariff": "УВЕРЕННЫЙ",
            "curator_id": str(curator_user.id),
            "csrf_token": "bypass",
        },
    )

    assert resp.status_code == 200
    student = db.query(User).filter(User.first_name == "Анна", User.last_name == "Иванова").first()
    assert student is not None
    assert student.vk_id < 0
    assert student.staff_login is not None
    assert student.password_hash is not None
    assert student.tg_username == "anna_arch"
    assert student.curator_id == curator_user.id
    assert student.profile_completed is False
    assert "Доступ создан" in resp.text


def test_superadmin_create_manual_student_reuses_existing_profile_by_tg_username(
    superadmin_client, db, user_factory, curator_user
):
    client, _ = superadmin_client
    existing = user_factory(
        vk_id=900105,
        name="Existing Student",
        role_name="ученик",
        profile_completed=True,
    )
    existing.first_name = "Мария"
    existing.last_name = "Соколова"
    existing.tg_username = "maria_arch"
    db.add(existing)
    db.commit()

    resp = client.post(
        "/cabinet/superadmin/users/create-student",
        data={
            "first_name": "Мария",
            "last_name": "Соколова",
            "tg_username": "@maria_arch",
            "tariff": "МАКСИМУМ",
            "curator_id": str(curator_user.id),
            "csrf_token": "bypass",
        },
    )

    assert resp.status_code == 200
    matches = [
        u for u in db.query(User).all()
        if (u.tg_username or "").lower() == "maria_arch"
    ]
    assert len(matches) == 1
    db.refresh(existing)
    assert existing.staff_login is not None
    assert existing.password_hash is not None
    assert existing.curator_id == curator_user.id
    assert existing.tariff == "МАКСИМУМ"
    assert existing.profile_completed is True


def test_superadmin_set_student_credentials_from_users_page(superadmin_client, db, user_factory):
    client, _ = superadmin_client
    student = user_factory(vk_id=900104, name="Manual Existing Student", role_name="ученик")

    resp = client.post(
        f"/cabinet/superadmin/users/{student.id}/set-credentials",
        data={"csrf_token": "bypass"},
    )

    assert resp.status_code == 200
    db.refresh(student)
    assert student.staff_login is not None
    assert student.password_hash is not None
    assert "Доступ создан" in resp.text


def test_superadmin_set_credentials_succeeds_for_admin_from_users_page(
    client, db, user_factory, session_factory
):
    admin = user_factory(vk_id=900109, name="Admin Actor", role_name="админ", is_admin=True)
    student = user_factory(vk_id=900110, name="Manual Existing Student 2", role_name="ученик")
    sess = session_factory(admin)
    client.cookies.set("session_id", sess.id)

    resp = client.post(
        f"/cabinet/superadmin/users/{student.id}/set-credentials",
        data={"csrf_token": "bypass"},
    )

    assert resp.status_code == 200
    db.refresh(student)
    assert student.staff_login is not None
    assert student.password_hash is not None
    assert "Доступ создан" in resp.text


def test_superadmin_set_credentials_denies_peer_rank_from_users_page(
    superadmin_client, db, user_factory
):
    client, _ = superadmin_client
    peer = user_factory(vk_id=900108, name="Peer Superadmin", role_name="суперадмин")

    resp = client.post(
        f"/cabinet/superadmin/users/{peer.id}/set-credentials",
        data={"csrf_token": "bypass"},
    )

    db.refresh(peer)
    assert resp.status_code == 403
    assert "Нельзя выдать доступ роли равной или выше своей" in resp.text
    assert peer.staff_login is None
    assert peer.password_hash is None


def test_superadmin_create_staff_from_users_page(superadmin_client, db):
    from app.models.role import Role

    client, _ = superadmin_client
    curator_role = db.query(Role).filter(Role.name == "куратор").first()

    resp = client.post(
        "/cabinet/superadmin/users/create-staff",
        data={
            "first_name": "Иван",
            "last_name": "Сотрудник",
            "role_id": str(curator_role.id),
            "csrf_token": "bypass",
        },
    )

    assert resp.status_code == 200
    staff = db.query(User).filter(User.first_name == "Иван", User.last_name == "Сотрудник").first()
    assert staff is not None
    assert staff.role_id == curator_role.id
    assert staff.vk_id < 0
    assert staff.staff_login is not None
    assert staff.password_hash is not None
    assert "Доступ создан" in resp.text


def test_superadmin_create_staff_denies_peer_rank_from_users_page(superadmin_client, db):
    from app.models.role import Role

    client, _ = superadmin_client
    superadmin_role = db.query(Role).filter(Role.name == "суперадмин").first()

    resp = client.post(
        "/cabinet/superadmin/users/create-staff",
        data={
            "first_name": "Пётр",
            "last_name": "Равный",
            "role_id": str(superadmin_role.id),
            "csrf_token": "bypass",
        },
    )

    assert resp.status_code == 200
    assert "Нельзя создать аккаунт с рангом не ниже вашего." in resp.text
    assert db.query(User).filter(User.first_name == "Пётр", User.last_name == "Равный").first() is None


def test_create_staff_page_denied_for_admin(client, db, user_factory, session_factory):
    """Создание сотрудника — единственное оставшееся отличие admin от superadmin."""
    admin = user_factory(vk_id=900111, name="Admin Actor", role_name="админ", is_admin=True)
    sess = session_factory(admin)
    client.cookies.set("session_id", sess.id)

    resp = client.get("/cabinet/superadmin/create-staff", follow_redirects=False)
    assert resp.status_code == 403


def test_create_staff_post_denied_for_admin(client, db, user_factory, session_factory):
    from app.models.role import Role

    admin = user_factory(vk_id=900112, name="Admin Actor", role_name="админ", is_admin=True)
    sess = session_factory(admin)
    client.cookies.set("session_id", sess.id)
    curator_role = db.query(Role).filter(Role.name == "куратор").first()

    resp = client.post(
        "/cabinet/superadmin/users/create-staff",
        data={
            "first_name": "Запрещено",
            "last_name": "Админу",
            "role_id": str(curator_role.id),
            "csrf_token": "bypass",
        },
        follow_redirects=False,
    )
    assert resp.status_code == 403
    assert db.query(User).filter(User.first_name == "Запрещено").first() is None


# ---------------------------------------------------------------------------
# GET /cabinet/superadmin/users — filters (migrated from test_routes_admin.py)
# ---------------------------------------------------------------------------

def test_superadmin_users_filter_by_role_rank(superadmin_client, db, user_factory):
    client, _ = superadmin_client
    user_factory(vk_id=900130, name="Role Filter Student", role_name="ученик")
    user_factory(vk_id=900131, name="Role Filter Curator", role_name="куратор")

    resp = client.get("/cabinet/superadmin/users?role_rank=2")

    assert resp.status_code == 200
    assert "Role Filter Curator" in resp.text
    assert "Role Filter Student" not in resp.text


def test_superadmin_users_filter_by_tag(superadmin_client, db, user_factory):
    from app.services.tags import add_tag_to_user, get_or_create_tag

    client, _ = superadmin_client
    tagged = user_factory(vk_id=900132, name="Tag Filter Tagged", role_name="ученик")
    other = user_factory(vk_id=900133, name="Tag Filter Other", role_name="ученик")

    tag = get_or_create_tag(db, "ОСОБЫЙ")
    add_tag_to_user(db, tagged.id, tag.id)
    db.commit()

    resp = client.get(f"/cabinet/superadmin/users?tag={tag.id}")

    assert resp.status_code == 200
    assert "Tag Filter Tagged" in resp.text
    assert "Tag Filter Other" not in resp.text


def test_superadmin_assign_role_from_users_page(superadmin_client, db, user_factory):
    from app.models.role import Role

    client, _ = superadmin_client
    target = user_factory(vk_id=900106, name="Role Change Student", role_name="ученик")
    curator_role = db.query(Role).filter(Role.name == "куратор").first()

    resp = client.post(
        f"/cabinet/superadmin/users/{target.id}/role",
        data={"role_id": str(curator_role.id), "csrf_token": "bypass"},
        follow_redirects=False,
    )

    assert resp.status_code == 303
    db.refresh(target)
    assert target.role_id == curator_role.id


def test_superadmin_role_cannot_change_own_role(superadmin_client, db):
    from app.models.role import Role

    client, admin = superadmin_client
    original_role_id = admin.role_id
    other_role = db.query(Role).filter(Role.id != admin.role_id).first()

    resp = client.post(
        f"/cabinet/superadmin/users/{admin.id}/role",
        data={"role_id": str(other_role.id), "csrf_token": "bypass"},
        follow_redirects=False,
    )

    assert resp.status_code == 303  # redirected, no-op (can_manage_user_by_rank rejects self)
    db.refresh(admin)
    assert admin.role_id == original_role_id


def test_superadmin_issue_link_from_users_page(superadmin_client, db, user_factory):
    client, _ = superadmin_client
    target = user_factory(vk_id=900107, name="Link Student", role_name="ученик", is_active=True)

    resp = client.post(
        f"/cabinet/superadmin/users/{target.id}/issue-link",
        data={"csrf_token": "bypass"},
    )

    assert resp.status_code == 200
    assert "/auth/link?token=" in resp.text
    token = db.query(LoginToken).filter(LoginToken.user_id == target.id).first()
    assert token is not None


def test_superadmin_issue_link_succeeds_for_admin_from_users_page(client, db, user_factory, session_factory):
    admin = user_factory(vk_id=900113, name="Admin Actor", role_name="админ", is_admin=True)
    target = user_factory(vk_id=900114, name="Link Student 2", role_name="ученик", is_active=True)
    sess = session_factory(admin)
    client.cookies.set("session_id", sess.id)

    resp = client.post(
        f"/cabinet/superadmin/users/{target.id}/issue-link",
        data={"csrf_token": "bypass"},
    )

    assert resp.status_code == 200
    assert "/auth/link?token=" in resp.text
    token = db.query(LoginToken).filter(LoginToken.user_id == target.id).first()
    assert token is not None


def test_superadmin_issue_link_denies_peer_rank_from_users_page(client, db, user_factory, session_factory):
    admin = user_factory(vk_id=900115, name="Admin Actor", role_name="админ", is_admin=True)
    peer = user_factory(vk_id=900116, name="Admin Peer", role_name="админ", is_active=True)
    sess = session_factory(admin)
    client.cookies.set("session_id", sess.id)

    resp = client.post(
        f"/cabinet/superadmin/users/{peer.id}/issue-link",
        data={"csrf_token": "bypass"},
        follow_redirects=False,
    )

    assert resp.status_code == 200
    assert "Нельзя выпустить ссылку для роли равной или выше своей." in resp.text
    token = db.query(LoginToken).filter(LoginToken.user_id == peer.id).first()
    assert token is None


# ---------------------------------------------------------------------------
# POST /cabinet/superadmin/issue-link
# ---------------------------------------------------------------------------

def test_issue_link_success(superadmin_client, db, curator_user):
    client, _ = superadmin_client
    resp = client.post("/cabinet/superadmin/issue-link",
                       data={"target_user_id": curator_user.id, "csrf_token": "bypass"})
    assert resp.status_code == 200

    # Token should be created in DB
    token = db.query(LoginToken).filter(LoginToken.user_id == curator_user.id).first()
    assert token is not None


def test_issue_link_shows_link_in_response(superadmin_client, db, curator_user):
    client, _ = superadmin_client
    resp = client.post("/cabinet/superadmin/issue-link",
                       data={"target_user_id": curator_user.id, "csrf_token": "bypass"})
    assert resp.status_code == 200
    assert "/auth/link?token=" in resp.text


def test_issue_link_nonexistent_user_404(superadmin_client):
    client, _ = superadmin_client
    resp = client.post("/cabinet/superadmin/issue-link",
                       data={"target_user_id": 99999, "csrf_token": "bypass"})
    assert resp.status_code == 404


def test_issue_link_succeeds_for_admin(client, db, user_factory, session_factory, curator_user):
    admin = user_factory(vk_id=900030, name="Admin", role_name="админ", is_admin=True)
    sess = session_factory(admin)
    client.cookies.set("session_id", sess.id)
    resp = client.post("/cabinet/superadmin/issue-link",
                       data={"target_user_id": curator_user.id, "csrf_token": "bypass"},
                       follow_redirects=False)
    assert resp.status_code == 200
    token = db.query(LoginToken).filter(LoginToken.user_id == curator_user.id).first()
    assert token is not None


def test_issue_link_denies_peer_rank_for_admin(client, db, user_factory, session_factory):
    admin = user_factory(vk_id=900031, name="Admin Actor", role_name="админ", is_admin=True)
    peer = user_factory(vk_id=900032, name="Admin Peer", role_name="админ")
    sess = session_factory(admin)
    client.cookies.set("session_id", sess.id)
    resp = client.post("/cabinet/superadmin/issue-link",
                       data={"target_user_id": peer.id, "csrf_token": "bypass"},
                       follow_redirects=False)
    assert resp.status_code == 403
    token = db.query(LoginToken).filter(LoginToken.user_id == peer.id).first()
    assert token is None


# ---------------------------------------------------------------------------
# POST /cabinet/superadmin/users/assign-curator-bulk
# ---------------------------------------------------------------------------

def test_bulk_assign_curator_by_tg_usernames(superadmin_client, db, user_factory, curator_user):
    client, _ = superadmin_client
    student_1 = user_factory(vk_id=900101, name="Student One")
    student_2 = user_factory(vk_id=900102, name="Student Two")
    student_1.tg_username = "student_one"
    student_2.tg_username = "student_two"
    db.add_all([student_1, student_2])
    db.commit()

    resp = client.post(
        "/cabinet/superadmin/users/assign-curator-bulk",
        data={
            "curator_id": str(curator_user.id),
            "student_usernames": "@student_one\nstudent_two\nmissing_user",
            "cohort_tag": "may",
            "csrf_token": "bypass",
        },
    )

    assert resp.status_code == 200
    db.refresh(student_1)
    db.refresh(student_2)
    assert student_1.curator_id == curator_user.id
    assert student_2.curator_id == curator_user.id
    assert student_1.cohort_tag == "may"
    assert student_2.cohort_tag == "may"
    assert "missing_user" in resp.text


def test_bulk_assign_curator_requires_superadmin(client, db, user_factory, session_factory, curator_user):
    student = user_factory(vk_id=900103, name="Plain Student")
    sess = session_factory(student)
    client.cookies.set("session_id", sess.id)

    resp = client.post(
        "/cabinet/superadmin/users/assign-curator-bulk",
        data={
            "curator_id": str(curator_user.id),
            "student_usernames": "student_one",
            "csrf_token": "bypass",
        },
        follow_redirects=False,
    )

    assert resp.status_code == 403


def test_superadmin_set_curator_ajax_returns_json(superadmin_client, db, user_factory, curator_user):
    client, _ = superadmin_client
    student = user_factory(vk_id=900203, name="Student Set Curator", role_name="ученик")

    resp = client.post(
        f"/cabinet/superadmin/users/{student.id}/curator",
        data={"curator_id": str(curator_user.id), "csrf_token": "bypass"},
    )

    db.refresh(student)
    data = resp.json()
    assert resp.status_code == 200
    assert data["ok"] is True
    assert data["user_id"] == student.id
    assert data["curator_id"] == curator_user.id
    assert data["curator_name"]
    assert student.curator_id == curator_user.id


def test_superadmin_set_cohort_tag_ajax_returns_json(superadmin_client, db, user_factory):
    client, _ = superadmin_client
    student = user_factory(vk_id=900210, name="Student Cohort Tag", role_name="ученик")

    resp = client.post(
        f"/cabinet/superadmin/users/{student.id}/cohort-tag",
        data={"cohort_tag": "june", "csrf_token": "bypass"},
    )

    db.refresh(student)
    data = resp.json()
    assert resp.status_code == 200
    assert data["ok"] is True
    assert data["user_id"] == student.id
    assert data["cohort_tag"] == "june"
    assert student.cohort_tag == "june"

    # Clearing the tag (empty value) resets it to None
    resp = client.post(
        f"/cabinet/superadmin/users/{student.id}/cohort-tag",
        data={"cohort_tag": "", "csrf_token": "bypass"},
    )
    db.refresh(student)
    data = resp.json()
    assert data["cohort_tag"] is None
    assert student.cohort_tag is None


def test_superadmin_set_cohort_tag_rejects_invalid_value(superadmin_client, db, user_factory):
    client, _ = superadmin_client
    student = user_factory(vk_id=900211, name="Student Bad Tag", role_name="ученик")

    resp = client.post(
        f"/cabinet/superadmin/users/{student.id}/cohort-tag",
        data={"cohort_tag": "winter", "csrf_token": "bypass"},
    )

    assert resp.status_code == 400


# ---------------------------------------------------------------------------
# POST /cabinet/superadmin/users/{id}/tags — tariff (migrated from test_routes_admin.py)
# ---------------------------------------------------------------------------

def test_superadmin_tags_update_changes_tariff(superadmin_client, db, user_factory):
    client, _ = superadmin_client
    target = user_factory(vk_id=900220, name="Tariff Student", tariff="УВЕРЕННЫЙ")

    resp = client.post(
        f"/cabinet/superadmin/users/{target.id}/tags",
        data={"tariff": "МАКСИМУМ", "csrf_token": "bypass"},
        follow_redirects=False,
    )

    assert resp.status_code == 303
    db.refresh(target)
    assert target.tariff == "МАКСИМУМ"


def test_superadmin_tags_update_invalid_tariff_rejected(superadmin_client, db, user_factory):
    client, _ = superadmin_client
    target = user_factory(vk_id=900221, name="Tariff Student Invalid", tariff="УВЕРЕННЫЙ")

    resp = client.post(
        f"/cabinet/superadmin/users/{target.id}/tags",
        data={"tariff": "INVALID_TARIFF", "csrf_token": "bypass"},
        follow_redirects=False,
    )

    assert resp.status_code == 400
    db.refresh(target)
    assert target.tariff == "УВЕРЕННЫЙ"  # unchanged


def test_bulk_assign_curator_rejects_inactive_curator(
    superadmin_client, db, user_factory
):
    client, _ = superadmin_client
    curator = user_factory(
        vk_id=900204,
        name="Inactive Curator",
        role_name="куратор",
        is_active=False,
    )
    student = user_factory(vk_id=900205, name="Student No Curator", role_name="ученик")
    student.tg_username = "student_no_curator"
    db.commit()

    resp = client.post(
        "/cabinet/superadmin/users/assign-curator-bulk",
        data={
            "curator_id": str(curator.id),
            "student_usernames": "student_no_curator",
            "csrf_token": "bypass",
        },
    )

    db.refresh(student)
    assert resp.status_code == 404
    assert student.curator_id is None


def test_superadmin_toggle_user_ajax_returns_json(superadmin_client, db, user_factory):
    client, _ = superadmin_client
    student = user_factory(vk_id=900201, name="Student Toggle")

    resp = client.post(
        f"/cabinet/superadmin/users/{student.id}/toggle-active",
        data={"csrf_token": "bypass"},
        headers={"Accept": "application/json", "X-Requested-With": "XMLHttpRequest"},
        follow_redirects=False,
    )

    assert resp.status_code == 200
    assert "location" not in resp.headers
    assert resp.json() == {"ok": True, "user_id": student.id, "is_active": False}
    db.refresh(student)
    assert student.is_active is False


def test_superadmin_toggle_user_re_enables(superadmin_client, db, user_factory):
    client, _ = superadmin_client
    student = user_factory(vk_id=900202, name="Student Re-enable", is_active=False)

    resp = client.post(
        f"/cabinet/superadmin/users/{student.id}/toggle-active",
        data={"csrf_token": "bypass"},
        headers={"Accept": "application/json", "X-Requested-With": "XMLHttpRequest"},
        follow_redirects=False,
    )

    assert resp.status_code == 200
    assert resp.json() == {"ok": True, "user_id": student.id, "is_active": True}
    db.refresh(student)
    assert student.is_active is True


def test_superadmin_delete_user_ajax_returns_json(superadmin_client, db, user_factory):
    client, _ = superadmin_client
    student = user_factory(vk_id=900202, name="Student Delete")

    resp = client.post(
        f"/cabinet/superadmin/users/{student.id}/delete",
        data={"csrf_token": "bypass"},
        headers={"Accept": "application/json", "X-Requested-With": "XMLHttpRequest"},
        follow_redirects=False,
    )

    assert resp.status_code == 200
    assert "location" not in resp.headers
    assert resp.json() == {"ok": True, "user_id": student.id, "deleted": True}
    db.refresh(student)
    assert student.deleted_at is not None
    assert student.is_active is False


# ---------------------------------------------------------------------------
# GET /cabinet/superadmin/stats — раздел «Полученные билеты»
# ---------------------------------------------------------------------------

def test_stats_page_shows_received_tickets(superadmin_client, db, user_factory):
    client, _ = superadmin_client
    student = user_factory(vk_id=900300, name="Тикет Ученик", role_name="ученик")
    db.add(MockExamAttempt(
        user_id=student.id,
        subject="Рисунок",
        ticket_id=1,
        ticket_title="Натюрморт с черепом",
        started_at=datetime(2026, 6, 14, 9, 0, tzinfo=timezone.utc),  # 12:00 MSK
    ))
    db.commit()

    resp = client.get("/cabinet/superadmin/stats")
    assert resp.status_code == 200
    assert "Полученные билеты" in resp.text
    assert "Натюрморт с черепом" in resp.text
    assert "14.06.2026 12:00" in resp.text   # MSK, не UTC 09:00


def test_stats_export_includes_ticket_sheets(superadmin_client, db, user_factory):
    client, _ = superadmin_client
    student = user_factory(vk_id=900301, name="Экспорт Ученик", role_name="ученик")
    db.add(MockExamAttempt(
        user_id=student.id,
        subject="Композиция",
        ticket_id=2,
        ticket_title="Билет для экспорта",
        started_at=datetime.now(timezone.utc),
    ))
    db.commit()

    resp = client.get("/cabinet/superadmin/stats/export")
    assert resp.status_code == 200
    assert resp.headers["content-type"].startswith(
        "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
    )

    import io
    import openpyxl
    wb = openpyxl.load_workbook(io.BytesIO(resp.content))
    assert "Билеты — сводка" in wb.sheetnames
    assert "Билеты — кто и когда" in wb.sheetnames
    titles = [row[1] for row in wb["Билеты — кто и когда"].iter_rows(min_row=2, values_only=True)]
    assert "Билет для экспорта" in titles


def _make_mock_with_feedback(db, student, curator):
    w = Work(
        user_id=student.id, work_type=WORK_TYPE_MOCK_EXAM, month="июнь", year=2026,
        filename="m.jpg", subject="Рисунок", score=77, scored_by_id=curator.id,
        status="success", is_final=True,
    )
    db.add(w)
    db.commit()
    db.refresh(w)
    fb = Feedback(work_id=w.id, curator_id=curator.id)
    db.add(fb)
    db.commit()
    db.refresh(fb)
    db.add(FeedbackMessage(
        feedback_id=fb.id, sender_id=curator.id, sender_role=ROLE_CURATOR,
        text="Композиция уравновешена, тон доработай",
    ))
    db.commit()
    return w


def test_stats_page_shows_feedback_table(superadmin_client, db, user_factory):
    client, _ = superadmin_client
    curator = user_factory(vk_id=900400, name="Куратор ОС", role_name="куратор")
    student = user_factory(vk_id=900401, name="ОС Ученик", role_name="ученик")
    _make_mock_with_feedback(db, student, curator)

    resp = client.get("/cabinet/superadmin/stats")
    assert resp.status_code == 200
    assert "Пробники и обратная связь" in resp.text
    assert "Композиция уравновешена" in resp.text


def test_stats_page_redesign_structure(superadmin_client, db, user_factory):
    """Редизайн: заголовок «Статистика пробников», секции в <details>,
    «Список сдач» удалён."""
    client, _ = superadmin_client
    resp = client.get("/cabinet/superadmin/stats")
    assert resp.status_code == 200
    assert "Статистика пробников" in resp.text
    assert "<details" in resp.text           # нативное сворачивание
    assert "<details open" not in resp.text  # все секции свёрнуты по умолчанию
    assert "Список сдач" not in resp.text     # убран полностью
    assert "Данные с 13.06.2026" in resp.text
    # все 5 секций присутствуют
    assert "Статус пробников по тарифам" in resp.text
    assert "Не сдали хотя бы один предмет" in resp.text
    assert "Полученные билеты" in resp.text
    assert "Пробники и обратная связь" in resp.text
    assert "Статистика по баллам" in resp.text


def test_stats_export_includes_feedback_sheet(superadmin_client, db, user_factory):
    client, _ = superadmin_client
    curator = user_factory(vk_id=900410, name="Куратор Эксп", role_name="куратор")
    student = user_factory(vk_id=900411, name="Эксп Ученик", role_name="ученик")
    _make_mock_with_feedback(db, student, curator)

    resp = client.get("/cabinet/superadmin/stats/export")
    assert resp.status_code == 200

    import io
    import openpyxl
    wb = openpyxl.load_workbook(io.BytesIO(resp.content))
    assert "Пробники + ОС" in wb.sheetnames
    rows = list(wb["Пробники + ОС"].iter_rows(min_row=2, values_only=True))
    assert any(r[0] and "Ученик" in str(r[0]) for r in rows)
    # столбец «Обратная связь» содержит текст ОС
    assert any(r[7] and "Композиция уравновешена" in str(r[7]) for r in rows)
