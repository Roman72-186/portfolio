import asyncio
import logging
import re
import secrets
import string
import time
import urllib.parse
import uuid
from datetime import datetime, timedelta, timezone, date
from typing import Annotated

logger = logging.getLogger(__name__)

import bcrypt as _bcrypt_lib
from fastapi import APIRouter, BackgroundTasks, Request, Depends, Form, HTTPException, Query, UploadFile, File
from fastapi.responses import HTMLResponse, JSONResponse, RedirectResponse
from sqlalchemy import func, case
from sqlalchemy.dialects.postgresql import insert as pg_insert
from sqlalchemy.orm import Session as DBSession, aliased

from app.config import settings
from app.constants import (
    MOCK_SUBJECTS,
    ASSIGNMENT_KINDS,
    ASSIGNMENT_KIND_LABELS,
    FEATURE_LABELS,
    FEATURE_PORTFOLIO_UPLOAD,
    FEATURE_MOCK_EXAM,
    FEATURE_RETAKE,
    TARIFFS,
    STUDY_MODES,
    STUDY_MODE_LABELS,
    EXAM_SUBJECT_HINTS,
    COHORT_TAGS,
    COHORT_TAG_LABELS,
)
from app.cache import invalidate_session as _invalidate_session_cache
from app.db.database import get_db
from app.dependencies import require_superadmin, require_admin_role, require_csrf, get_current_user
from app.models.exam_assignment import ExamAssignment, ExamTicket, ExamTicketAssignee
from app.models.exam_cycle import ExamCycle
from app.models.feature_period import FeaturePeriod
from app.services.feature_periods import invalidate_feature_cache, get_active_period
from app.services.tz import MSK_TZ, now_msk, today_msk, msk_midnight
from app.models.role import Role
from app.models.tag import Tag, UserTag
from app.models.user import User
from app.models.work import Work, WORK_TYPE_MOCK_EXAM
from app.services import s3 as s3_service
from app.services.auth_links import issue_one_time_login_link, issue_telegram_link_token, next_manual_vk_id
from app.services.tags import get_all_tags
from app.services.mock_exam_access import (
    MOCK_EXAM_DEFAULT_DURATION_MINUTES,
    ticket_closes_at,
    ticket_duration_sec,
    ticket_opens_at,
    ticket_start_cutoff_at,
)
from app.services.utils import compress_image, rotate_image_bytes
from app.tmpl import templates

_TRANSLIT = str.maketrans(
    "абвгдеёжзийклмнопрстуфхцчшщъыьэюяАБВГДЕЁЖЗИЙКЛМНОПРСТУФХЦЧШЩЪЫЬЭЮЯ",
    "abvgdeejzijklmnoprstufhccssxyxeuaABVGDEEJZIJKLMNOPRSTUFHCCSSXYXEUA",
)
_PWD_CHARS = "abcdefghjkmnpqrstuvwxyz23456789"
_TG_USERNAME_RE = re.compile(r"^[A-Za-z0-9_]{4,32}$")


def _transliterate(s: str) -> str:
    return s.translate(_TRANSLIT).lower()


def _gen_password(length: int = 10) -> str:
    return "".join(secrets.choice(_PWD_CHARS) for _ in range(length))


def _hash_password(plain: str) -> str:
    return _bcrypt_lib.hashpw(plain.encode(), _bcrypt_lib.gensalt()).decode()


def _make_login(user: User, db: DBSession) -> str:
    """Generate a unique staff_login like 'ivan.s' based on first/last name."""
    first = _transliterate(user.first_name or user.name or "user")
    last_initial = _transliterate((user.last_name or "")[:1])
    base = f"{first}.{last_initial}" if last_initial else first
    # Keep only safe ascii chars
    base = "".join(c for c in base if c in string.ascii_lowercase + string.digits + ".")
    base = base[:20] or "user"

    candidate = base
    suffix = 2
    while db.query(User).filter(User.staff_login == candidate, User.id != user.id).first():
        candidate = f"{base}{suffix}"
        suffix += 1
    return candidate


def _display_user_name(user: User) -> str:
    return f"{user.last_name or ''} {user.first_name or user.name}".strip()


def _issue_login_password(db: DBSession, target: User) -> dict:
    """Generate or reset password credentials for manual login via /login."""
    if not target.staff_login:
        target.staff_login = _make_login(target, db)

    new_password = _gen_password()
    target.password_hash = _hash_password(new_password)
    return {
        "name": _display_user_name(target),
        "login": target.staff_login,
        "password": new_password,
        "url": f"https://{settings.domain}/login" if settings.domain else "/login",
    }


router = APIRouter(prefix="/cabinet")


# ── Поворот фото в S3 (только суперадмин) ────────────────────────────────────
#
# Суперадмин при просмотре в лайтбоксе может повернуть любое фото на 90°. Поворот
# деструктивный: скачиваем объект из S3, крутим на полном разрешении и
# перезаписываем тот же ключ (s3_url/s3_path не меняются → запись Work не трогаем).
# Авторизация на конкретный файл — через s3_path_from_public_url: он вернёт None
# для любой ссылки вне нашего бакета. Видео (отчёты кураторов) отсекаются тем, что
# PIL не сможет открыть их как изображение.
@router.post("/rotate-photo")
async def rotate_photo(
    user: Annotated[dict, Depends(require_superadmin)],
    _csrf: Annotated[None, Depends(require_csrf)],
    src: Annotated[str, Form()],
    direction: Annotated[str, Form()],
):
    if direction not in ("left", "right"):
        return JSONResponse({"success": False, "error": "Неверное направление поворота"}, status_code=422)
    if not s3_service.is_configured():
        return JSONResponse({"success": False, "error": "S3 не настроен"}, status_code=503)

    # Срезаем cache-busting ?v=… (повторный поворот шлёт уже изменённый URL).
    clean_url = src.split("?", 1)[0]
    s3_path = s3_service.s3_path_from_public_url(clean_url)
    if not s3_path:
        return JSONResponse({"success": False, "error": "Неизвестный файл"}, status_code=400)
    # Браузер отдаёт img.src percent-encoded (пути пробников/портфолио содержат
    # кириллицу: тариф, папки «До»/«После»). Ключ в S3 — сырой, поэтому декодируем.
    s3_path = urllib.parse.unquote(s3_path)

    def _do() -> tuple[str | None, str | None]:
        data = s3_service.download_from_s3(s3_path)
        if data is None:
            return None, "Не удалось загрузить файл из хранилища"
        try:
            rotated = rotate_image_bytes(data, clockwise=(direction == "right"))
        except Exception:  # noqa: BLE001 — PIL не открыл (видео/битый файл)
            return None, "Это не изображение — поворот недоступен"
        new_url = s3_service.upload_to_s3(s3_path, rotated, "image/jpeg")
        if not new_url:
            return None, "Не удалось сохранить повёрнутое фото"
        return new_url, None

    loop = asyncio.get_running_loop()
    new_url, err = await loop.run_in_executor(None, _do)
    if err:
        return JSONResponse({"success": False, "error": err}, status_code=422)

    logger.info("rotate-photo by %s: %s (%s)", user.get("user_id"), s3_path, direction)
    return JSONResponse({"success": True, "src": f"{new_url}?v={int(time.time())}"})


def _month_name_prep(month: int) -> str:
    names = ["", "январе", "феврале", "марте", "апреле", "мае", "июне",
             "июле", "августе", "сентябре", "октябре", "ноябре", "декабре"]
    return names[month] if 1 <= month <= 12 else ""


def _load_dashboard_data(db: DBSession, now: datetime) -> dict:
    month_start = datetime(now.year, now.month, 1, tzinfo=timezone.utc)

    # ── Users by role ─────────────────────────────────────────────────────────
    role_rows = (
        db.query(Role.display_name, Role.rank, func.count(User.id).label("cnt"))
        .outerjoin(User, (User.role_id == Role.id) & (User.is_active == True))
        .group_by(Role.id, Role.display_name, Role.rank)
        .order_by(Role.rank)
        .all()
    )
    role_breakdown = [{"name": r.display_name, "rank": r.rank, "count": r.cnt} for r in role_rows]
    total_active = sum(r["count"] for r in role_breakdown)
    inactive_count = db.query(func.count(User.id)).filter(User.is_active == False).scalar() or 0
    new_users_month = (
        db.query(func.count(User.id)).filter(User.created_at >= month_start).scalar() or 0
    )

    # ── Works ─────────────────────────────────────────────────────────────────
    works_type_rows = (
        db.query(Work.work_type, func.count(Work.id))
        .filter(Work.status == "success")
        .group_by(Work.work_type)
        .all()
    )
    works_by_type = {wt: cnt for wt, cnt in works_type_rows}
    total_works = sum(works_by_type.values())
    works_this_month = (
        db.query(func.count(Work.id))
        .filter(Work.status == "success", Work.created_at >= month_start)
        .scalar() or 0
    )

    # ── Scores ────────────────────────────────────────────────────────────────
    avg_score_raw = (
        db.query(func.avg(Work.score))
        .filter(Work.score.isnot(None), Work.status == "success")
        .scalar()
    )
    avg_score = round(float(avg_score_raw)) if avg_score_raw is not None else None
    mock_period = get_active_period(db, FEATURE_MOCK_EXAM)
    if mock_period:
        _mock_start = msk_midnight(mock_period.start_date)
        _mock_end = msk_midnight(mock_period.end_date + timedelta(days=1))
        unscored_mocks = (
            db.query(func.count(Work.id))
            .filter(
                Work.work_type == WORK_TYPE_MOCK_EXAM,
                Work.score.is_(None),
                Work.status == "success",
                Work.created_at >= _mock_start,
                Work.created_at < _mock_end,
            )
            .scalar() or 0
        )
    else:
        unscored_mocks = 0

    # ── Curators (rank 2) ─────────────────────────────────────────────────────
    # Limit: max 100 curators (защита от медленной выборки)
    StudentAlias = aliased(User)
    curator_rows = (
        db.query(
            User.id, User.first_name, User.last_name, User.name, User.photo_url,
            func.count(StudentAlias.id).label("student_count"),
        )
        .join(Role, User.role_id == Role.id)
        .outerjoin(StudentAlias, (StudentAlias.curator_id == User.id) & (StudentAlias.is_active == True))
        .filter(Role.rank == 2, User.is_active == True)
        .group_by(User.id, User.first_name, User.last_name, User.name, User.photo_url)
        .order_by(func.count(StudentAlias.id).desc())
        .limit(100)
        .all()
    )
    curators = [
        {
            "id": r.id,
            "name": f"{r.last_name or ''} {r.first_name or r.name}".strip(),
            "photo_url": r.photo_url,
            "student_count": r.student_count,
        }
        for r in curator_rows
    ]

    # ── Admins (rank 4) ───────────────────────────────────────────────────────
    # Limit: max 50 admins (защита от медленной выборки)
    admin_rows = (
        db.query(User.id, User.first_name, User.last_name, User.name, User.photo_url)
        .join(Role, User.role_id == Role.id)
        .filter(Role.rank == 4, User.is_active == True)
        .order_by(User.last_name, User.first_name)
        .limit(50)
        .all()
    )
    admins = [
        {
            "id": r.id,
            "name": f"{r.last_name or ''} {r.first_name or r.name}".strip(),
            "photo_url": r.photo_url,
        }
        for r in admin_rows
    ]

    # ── Recent uploads ────────────────────────────────────────────────────────
    recent_rows = (
        db.query(Work, User)
        .join(User, Work.user_id == User.id)
        .filter(Work.status == "success")
        .order_by(Work.created_at.desc())
        .limit(10)
        .all()
    )
    recent_works = [
        {
            "work_type": w.work_type,
            "filename": w.filename,
            "created_at": w.created_at,
            "s3_url": w.s3_url,
            "student_name": f"{u.last_name or ''} {u.first_name or u.name}".strip(),
            "student_id": u.id,
            "score": float(w.score) if w.score is not None else None,
        }
        for w, u in recent_rows
    ]

    # ── Feature periods status ────────────────────────────────────────────────
    today = today_msk()
    active_features = set(
        row[0]
        for row in db.query(FeaturePeriod.feature)
        .filter(
            FeaturePeriod.is_active.is_(True),
            FeaturePeriod.start_date <= today,
            FeaturePeriod.end_date >= today,
        )
        .all()
    )
    feature_statuses = {
        feat: {"label": FEATURE_LABELS[feat], "open": feat in active_features}
        for feat in ALL_FEATURES
    }

    return {
        "total_active": total_active,
        "inactive_count": inactive_count,
        "new_users_month": new_users_month,
        "role_breakdown": role_breakdown,
        "total_works": total_works,
        "works_this_month": works_this_month,
        "works_by_type": works_by_type,
        "avg_score": avg_score,
        "unscored_mocks": unscored_mocks,
        "curators": curators,
        "admins": admins,
        "recent_works": recent_works,
        "month_name": _month_name_prep(now.month),
        "feature_statuses": feature_statuses,
    }


@router.get("/superadmin", response_class=HTMLResponse)
def cabinet_superadmin(
    request: Request,
    user: Annotated[dict, Depends(require_admin_role)],
    db: Annotated[DBSession, Depends(get_db)],
):
    now = datetime.now(timezone.utc)
    ctx = _load_dashboard_data(db, now)
    ctx.update({"request": request, "user": user})
    return templates.TemplateResponse("cabinet_staff.html", ctx)


@router.post("/superadmin/set-credentials", response_class=HTMLResponse)
def superadmin_set_credentials(
    request: Request,
    user: Annotated[dict, Depends(require_admin_role)],
    db: Annotated[DBSession, Depends(get_db)],
    _csrf: Annotated[None, Depends(require_csrf)],
    target_user_id: int = Form(...),
):
    target = db.query(User).filter(User.id == target_user_id, User.is_active == True).first()
    if not target:
        raise HTTPException(status_code=404, detail="Пользователь не найден")
    if not can_manage_user_by_rank(user["user_id"], user["role_rank"], target):
        raise HTTPException(status_code=403, detail="Нельзя выдать доступ роли равной или выше своей")

    issued_creds = _issue_login_password(db, target)
    db.commit()

    now = datetime.now(timezone.utc)
    ctx = _load_dashboard_data(db, now)
    ctx.update({
        "request": request,
        "user": user,
        "issued_creds": issued_creds,
    })
    return templates.TemplateResponse("cabinet_staff.html", ctx)


@router.post("/superadmin/issue-link", response_class=HTMLResponse)
def superadmin_issue_link(
    request: Request,
    user: Annotated[dict, Depends(require_admin_role)],
    db: Annotated[DBSession, Depends(get_db)],
    _csrf: Annotated[None, Depends(require_csrf)],
    target_user_id: int = Form(...),
):
    target = db.query(User).filter(User.id == target_user_id, User.is_active == True).first()
    if not target:
        raise HTTPException(status_code=404, detail="Пользователь не найден")
    if not can_manage_user_by_rank(user["user_id"], user["role_rank"], target):
        raise HTTPException(status_code=403, detail="Нельзя выдать доступ роли равной или выше своей")

    base_url = f"https://{settings.domain}" if settings.domain else str(request.base_url).rstrip("/")
    issued_link, login_token = issue_one_time_login_link(
        db,
        user=target,
        base_url=base_url,
        issued_by=f"superadmin:{user['user_id']}",
    )

    now = datetime.now(timezone.utc)
    ctx = _load_dashboard_data(db, now)
    ctx.update({
        "request": request,
        "user": user,
        "issued_link": issued_link,
        "issued_link_name": f"{target.last_name or ''} {target.first_name or target.name}".strip(),
        "issued_link_expires_at": login_token.expires_at,
    })
    return templates.TemplateResponse("cabinet_staff.html", ctx)


# ═══════════════════════════════════════════════════════════════════════════════
# Exam Assignments — создание заданий для сдачи пробников
# ═══════════════════════════════════════════════════════════════════════════════

def _ticket_s3_path(filename: str) -> str:
    ext = filename.rsplit(".", 1)[-1].lower() if "." in filename else "jpg"
    key = uuid.uuid4().hex[:12]
    return f"Экзамены/билеты/{key}.{ext}"


# ── Редирект старых URL (backward compat) ────────────────────────────────────

@router.get("/superadmin/exam-assignments")
def _exam_assignments_compat(_user: Annotated[dict, Depends(require_admin_role)]):
    return RedirectResponse("/cabinet/exam-assignments", status_code=301)

@router.get("/superadmin/exam-assignments/create")
def _exam_assignment_create_compat(_user: Annotated[dict, Depends(require_admin_role)]):
    return RedirectResponse("/cabinet/exam-assignments/create", status_code=301)

@router.get("/superadmin/exam-assignments/{assignment_id}")
def _exam_assignment_detail_compat(assignment_id: int, _user: Annotated[dict, Depends(require_admin_role)]):
    return RedirectResponse(f"/cabinet/exam-assignments/{assignment_id}", status_code=301)


# ── Хаб «Билеты для пробников» ───────────────────────────────────────────────

@router.get("/exam-assignments", response_class=HTMLResponse)
def exam_assignments_hub(
    request: Request,
    user: Annotated[dict, Depends(require_admin_role)],
    db: Annotated[DBSession, Depends(get_db)],
):
    counts = dict(
        db.query(ExamAssignment.status, func.count(ExamAssignment.id))
        .filter(ExamAssignment.kind != "guest")
        .group_by(ExamAssignment.status)
        .all()
    )
    return templates.TemplateResponse("superadmin_exam_hub.html", {
        "request": request,
        "user": user,
        "active_count": counts.get("published", 0) + counts.get("draft", 0),
        "archived_count": counts.get("archived", 0),
    })


def _render_assignment_list(request, user, db: DBSession, statuses: list[str], mode: str):
    """Общий рендер списка заданий для вкладок «Активные» (published+draft) и «Архив»."""
    assignments = (
        db.query(ExamAssignment)
        .filter(ExamAssignment.status.in_(statuses), ExamAssignment.kind != "guest")
        # published сверху, остальные — по дате создания (свежие выше)
        .order_by((ExamAssignment.status != "published").asc(), ExamAssignment.created_at.desc())
        .limit(200)
        .all()
    )
    ticket_counts: dict[int, int] = {}
    period_by_assignment: dict[int, tuple] = {}
    if assignments:
        rows = (
            db.query(
                ExamTicket.assignment_id,
                func.count(ExamTicket.id),
                func.min(ExamTicket.start_date),
                func.max(ExamTicket.end_date),
            )
            .filter(ExamTicket.assignment_id.in_([a.id for a in assignments]))
            .group_by(ExamTicket.assignment_id)
            .all()
        )
        for aid, cnt, start_min, end_max in rows:
            ticket_counts[aid] = cnt
            period_by_assignment[aid] = (start_min, end_max)

    return templates.TemplateResponse("superadmin_exam_assignments.html", {
        "request": request,
        "user": user,
        "assignments": assignments,
        "ticket_counts": ticket_counts,
        "period_by_assignment": period_by_assignment,
        "mode": mode,
    })


@router.get("/exam-assignments/active", response_class=HTMLResponse)
def exam_assignments_active(
    request: Request,
    user: Annotated[dict, Depends(require_admin_role)],
    db: Annotated[DBSession, Depends(get_db)],
):
    return _render_assignment_list(request, user, db, ["published", "draft"], "active")


@router.get("/exam-assignments/archive", response_class=HTMLResponse)
def exam_assignments_archive_list(
    request: Request,
    user: Annotated[dict, Depends(require_admin_role)],
    db: Annotated[DBSession, Depends(get_db)],
):
    return _render_assignment_list(request, user, db, ["archived"], "archive")


# ── Форма создания ───────────────────────────────────────────────────────────

def _load_student_list(db: DBSession) -> list[dict]:
    student_role = db.query(Role).filter(Role.rank == 1).first()
    if not student_role:
        return []
    students = (
        db.query(
            User.id, User.first_name, User.last_name, User.name, User.tg_username,
            User.curator_id, User.tariff, User.cohort_tag, User.study_mode,
        )
        .filter(User.role_id == student_role.id, User.is_active == True)
        .order_by(User.last_name, User.first_name)
        .all()
    )
    # Теги учеников (M2M) одним запросом — для конструктора подбора получателей.
    tags_by_user: dict[int, list[int]] = {}
    student_ids = [s.id for s in students]
    if student_ids:
        for uid, tid in (
            db.query(UserTag.user_id, UserTag.tag_id)
            .filter(UserTag.user_id.in_(student_ids))
            .all()
        ):
            tags_by_user.setdefault(uid, []).append(tid)
    return [
        {
            "id": s.id,
            "name": f"{s.last_name or ''} {s.first_name or s.name}".strip(),
            "username": (s.tg_username or "").strip().lstrip("@"),
            "curator_id": s.curator_id,
            "tariff": s.tariff or "",
            "cohort_tag": s.cohort_tag or "",
            "study_mode": s.study_mode or "",
            "tag_ids": tags_by_user.get(s.id, []),
        }
        for s in students
    ]


def _load_curator_list(db: DBSession) -> list[dict]:
    """Активные кураторы (rank=2) для фильтра подбора получателей."""
    curator_role = db.query(Role).filter(Role.rank == 2).first()
    if not curator_role:
        return []
    curators = (
        db.query(User.id, User.first_name, User.last_name, User.name)
        .filter(User.role_id == curator_role.id, User.is_active == True)
        .order_by(User.last_name, User.first_name)
        .all()
    )
    return [
        {"id": c.id, "name": f"{c.last_name or ''} {c.first_name or c.name}".strip()}
        for c in curators
    ]


def _load_tag_list(db: DBSession) -> list[dict]:
    return [{"id": tag.id, "name": tag.name} for tag in get_all_tags(db)]


def _next_seq_number(db: DBSession, kind: str, subject: str) -> int:
    """Сквозной порядковый номер задания в пределах (kind, subject): MAX+1.

    Монотонный: удалённые номера не переиспользуются. Считается только при create.
    """
    current = (
        db.query(func.max(ExamAssignment.seq_number))
        .filter(ExamAssignment.kind == kind, ExamAssignment.subject == subject)
        .scalar()
    )
    return (current or 0) + 1


def _compose_assignment_title(
    kind: str, seq: int | None, subject: str, created: date, note: str | None
) -> str:
    """Авто-название: «Пробник №5 · Рисунок · 22.06.2026 · примечание»."""
    label = ASSIGNMENT_KIND_LABELS.get(kind, ASSIGNMENT_KIND_LABELS["mock"])
    head = f"{label} №{seq}" if seq else label
    parts = [head, subject, created.strftime("%d.%m.%Y")]
    if note:
        parts.append(note)
    return " · ".join(parts)


def _default_ticket_schedule() -> dict:
    base = now_msk()
    if base.time() >= datetime.strptime("18:30", "%H:%M").time():
        base = base + timedelta(days=1)
    open_at = base.replace(hour=11, minute=45, second=0, microsecond=0)
    close_at = base.replace(hour=18, minute=30, second=0, microsecond=0)
    return {
        "opens_at": open_at.strftime("%Y-%m-%dT%H:%M"),
        "closes_at": close_at.strftime("%Y-%m-%dT%H:%M"),
        "duration_minutes": MOCK_EXAM_DEFAULT_DURATION_MINUTES,
    }


def _parse_msk_datetime_local(raw: str, *, ticket_number: int, field_label: str) -> datetime:
    try:
        value = datetime.fromisoformat(raw)
    except ValueError:
        raise HTTPException(
            status_code=422,
            detail=f"Билет {ticket_number}: неверное время «{field_label}»",
        )
    if value.tzinfo is None:
        value = value.replace(tzinfo=MSK_TZ)
    else:
        value = value.astimezone(MSK_TZ)
    return value.astimezone(timezone.utc)


def _format_msk_datetime_local(value: datetime) -> str:
    if value.tzinfo is None:
        value = value.replace(tzinfo=timezone.utc)
    return value.astimezone(MSK_TZ).strftime("%Y-%m-%dT%H:%M")


def _format_msk_display(value: datetime) -> str:
    if value.tzinfo is None:
        value = value.replace(tzinfo=timezone.utc)
    return value.astimezone(MSK_TZ).strftime("%d.%m.%Y %H:%M")


@router.get("/exam-assignments/create", response_class=HTMLResponse)
def exam_assignment_create_form(
    request: Request,
    user: Annotated[dict, Depends(require_admin_role)],
    db: Annotated[DBSession, Depends(get_db)],
):
    return templates.TemplateResponse("superadmin_exam_assignment_form.html", {
        "request": request,
        "user": user,
        "subjects": MOCK_SUBJECTS,
        "kind_labels": ASSIGNMENT_KIND_LABELS,
        "student_list": _load_student_list(db),
        "tag_list": _load_tag_list(db),
        "curator_list": _load_curator_list(db),
        "tariffs": TARIFFS,
        "cohort_labels": COHORT_TAG_LABELS,
        "study_mode_labels": STUDY_MODE_LABELS,
        "default_schedule": _default_ticket_schedule(),
        "is_edit": False,
    })


@router.get("/exam-assignments/{assignment_id}/edit", response_class=HTMLResponse)
def exam_assignment_edit_form(
    assignment_id: int,
    request: Request,
    user: Annotated[dict, Depends(require_admin_role)],
    db: Annotated[DBSession, Depends(get_db)],
):
    assignment = db.query(ExamAssignment).filter(ExamAssignment.id == assignment_id).first()
    if not assignment:
        raise HTTPException(status_code=404, detail="Задание не найдено")

    tickets = (
        db.query(ExamTicket)
        .filter(ExamTicket.assignment_id == assignment_id)
        .order_by(ExamTicket.ticket_number)
        .all()
    )

    # Load assignees per ticket
    assignees_by_ticket: dict[int, list[int]] = {t.id: [] for t in tickets}
    if tickets:
        rows = (
            db.query(ExamTicketAssignee.ticket_id, ExamTicketAssignee.user_id)
            .filter(ExamTicketAssignee.ticket_id.in_([t.id for t in tickets]))
            .all()
        )
        for ticket_id, user_id in rows:
            assignees_by_ticket.setdefault(ticket_id, []).append(user_id)

    existing_tickets = [
        {
            "title": t.title,
            "description": t.description or "",
            "image_url": t.image_s3_url or "",
            "image_path": t.image_s3_path or "",
            "start_date": t.start_date.isoformat() if t.start_date else "",
            "end_date": t.end_date.isoformat() if t.end_date else "",
            "opens_at": _format_msk_datetime_local(ticket_opens_at(t)),
            "closes_at": _format_msk_datetime_local(ticket_closes_at(t)),
            "duration_minutes": ticket_duration_sec(t) // 60,
            "restrict_start_by_duration": bool(t.restrict_start_by_duration),
            "target_tag_id": t.target_tag_id,
            "assign_to_all": bool(t.assign_to_all),
            "student_ids": assignees_by_ticket.get(t.id, []),
        }
        for t in tickets
    ]

    assignment_date = (
        assignment.created_at.astimezone(MSK_TZ).date()
        if assignment.created_at else today_msk()
    )

    import json as _json
    return templates.TemplateResponse("superadmin_exam_assignment_form.html", {
        "request": request,
        "user": user,
        "subjects": MOCK_SUBJECTS,
        "kind_labels": ASSIGNMENT_KIND_LABELS,
        "student_list": _load_student_list(db),
        "tag_list": _load_tag_list(db),
        "curator_list": _load_curator_list(db),
        "tariffs": TARIFFS,
        "cohort_labels": COHORT_TAG_LABELS,
        "study_mode_labels": STUDY_MODE_LABELS,
        "default_schedule": _default_ticket_schedule(),
        "is_edit": True,
        "assignment": assignment,
        "assignment_date_str": assignment_date.strftime("%d.%m.%Y"),
        "existing_tickets_json": _json.dumps(existing_tickets, ensure_ascii=False),
    })


# ── AJAX загрузка фото билета ────────────────────────────────────────────────

@router.post("/upload-ticket-image")
async def upload_ticket_image(
    user: Annotated[dict, Depends(require_admin_role)],
    _csrf: Annotated[None, Depends(require_csrf)],
    file: UploadFile = File(...),
):
    ct = (file.content_type or "").lower()
    fname = file.filename or "image.jpg"
    ext = ("." + fname.rsplit(".", 1)[-1].lower()) if "." in fname else ""
    allowed_ext = {".jpg", ".jpeg", ".png", ".webp", ".heic", ".heif", ".avif", ".gif", ".bmp", ".tif", ".tiff"}
    if not ct.startswith("image/") and ext not in allowed_ext:
        return JSONResponse({"success": False, "error": "Файл не является изображением"}, status_code=422)

    data = await file.read()
    if len(data) > 10 * 1024 * 1024:
        return JSONResponse({"success": False, "error": "Файл слишком большой (макс. 10 МБ)"}, status_code=413)
    if not data:
        return JSONResponse({"success": False, "error": "Пустой файл"}, status_code=422)

    compressed = compress_image(data)
    s3_path = _ticket_s3_path(fname)
    url = s3_service.upload_to_s3(s3_path, compressed, "image/jpeg")
    if s3_service.is_configured() and not url:
        return JSONResponse({"success": False, "error": "Ошибка загрузки в хранилище"}, status_code=502)
    return JSONResponse({"success": True, "url": url, "path": s3_path if url else None})


# ── Сохранение задания ───────────────────────────────────────────────────────

@router.post("/exam-assignments/create")
async def exam_assignment_create_submit(
    request: Request,
    user: Annotated[dict, Depends(require_admin_role)],
    db: Annotated[DBSession, Depends(get_db)],
    _csrf: Annotated[None, Depends(require_csrf)],
):
    form = await request.form()

    kind = str(form.get("kind", "mock")).strip()
    subject = str(form.get("subject", "")).strip()
    note = str(form.get("note", "")).strip() or None
    ticket_count = int(form.get("ticket_count", 1) or 1)
    ticket_count = max(1, min(10, ticket_count))

    if kind not in ASSIGNMENT_KINDS:
        raise HTTPException(status_code=422, detail="Неверный тип задания")
    if subject not in MOCK_SUBJECTS:
        raise HTTPException(status_code=422, detail="Неверный предмет")

    # Авто-номер (сквозной по kind+subject) и авто-название с датой создания.
    seq_number = _next_seq_number(db, kind, subject)
    title = _compose_assignment_title(kind, seq_number, subject, today_msk(), note)

    assignment = ExamAssignment(
        title=title,
        subject=subject,
        kind=kind,
        seq_number=seq_number,
        note=note,
        created_by_id=user["user_id"],
        status="published",
    )
    db.add(assignment)
    db.flush()

    _build_tickets_from_form(db, assignment, form, ticket_count)
    _ensure_mock_exam_period_open(db, user["user_id"])
    db.commit()
    return RedirectResponse(
        f"/cabinet/exam-assignments/{assignment.id}", status_code=303
    )


def _ensure_mock_exam_period_open(db: DBSession, created_by_id: int) -> None:
    """Гарантирует наличие активного FeaturePeriod для пробников от сегодня
    до самой поздней end_date билетов в БД. Если такого нет — создаёт.

    Решает usability-проблему: главный преподаватель создал билет, но забыл открыть период
    в /cabinet/periods → ученики ничего не видели.
    """
    today = today_msk()
    max_end_row = (
        db.query(func.max(ExamTicket.end_date))
        .join(ExamAssignment, ExamTicket.assignment_id == ExamAssignment.id)
        .filter(ExamAssignment.status == "published", ExamTicket.end_date >= today)
        .first()
    )
    max_end = max_end_row[0] if max_end_row else None
    if not max_end:
        return  # нечего открывать

    # Уже есть активный период, покрывающий today?
    existing = (
        db.query(FeaturePeriod)
        .filter(
            FeaturePeriod.feature == FEATURE_MOCK_EXAM,
            FeaturePeriod.is_active.is_(True),
            FeaturePeriod.start_date <= today,
            FeaturePeriod.end_date >= today,
        )
        .first()
    )
    if existing:
        # Если max_end позднее текущего end_date — расширяем
        if existing.end_date < max_end:
            existing.end_date = max_end
            db.flush()
            invalidate_feature_cache(FEATURE_MOCK_EXAM)
        return

    # Создаём новый период
    db.add(FeaturePeriod(
        feature=FEATURE_MOCK_EXAM,
        title="Авто-открыто при публикации билета",
        start_date=today,
        end_date=max_end,
        is_active=True,
        created_by_id=created_by_id,
    ))
    db.flush()
    invalidate_feature_cache(FEATURE_MOCK_EXAM)


def _build_tickets_from_form(db: DBSession, assignment, form, ticket_count: int) -> None:
    """Парсит ticket_{i}_* поля формы и создаёт ExamTicket + ExamTicketAssignee.
    Используется и при создании, и при редактировании (после удаления старых тикетов).
    """
    for i in range(1, ticket_count + 1):
        t_title = str(form.get(f"ticket_{i}_title", "")).strip()
        t_desc = str(form.get(f"ticket_{i}_description", "")).strip() or None
        t_img_url = str(form.get(f"ticket_{i}_image_url", "")).strip() or None
        t_img_path = str(form.get(f"ticket_{i}_image_path", "")).strip() or None
        t_opens_raw = str(form.get(f"ticket_{i}_opens_at", "")).strip()
        t_closes_raw = str(form.get(f"ticket_{i}_closes_at", "")).strip()
        t_duration_raw = str(form.get(f"ticket_{i}_duration_minutes", "")).strip()
        restrict_start_by_duration = form.get(f"ticket_{i}_restrict_start_by_duration") == "on"
        t_target_tag_raw = str(form.get(f"ticket_{i}_target_tag_id", "")).strip()
        t_activate_mode = str(form.get(f"ticket_{i}_activate_mode", "scheduled")).strip()
        t_start_raw = str(form.get(f"ticket_{i}_start_date", "")).strip()
        t_end_raw = str(form.get(f"ticket_{i}_end_date", "")).strip()
        t_all = form.get(f"ticket_{i}_assign_all") == "on"
        t_students_raw = str(form.get(f"ticket_{i}_student_ids", "")).strip()

        if not t_title:
            raise HTTPException(status_code=422, detail=f"Название билета {i} обязательно")

        if t_duration_raw:
            try:
                duration_minutes = int(t_duration_raw)
            except ValueError:
                raise HTTPException(status_code=422, detail=f"Билет {i}: неверный таймер сдачи")
        else:
            duration_minutes = MOCK_EXAM_DEFAULT_DURATION_MINUTES
        if duration_minutes < 1 or duration_minutes > 720:
            raise HTTPException(
                status_code=422,
                detail=f"Билет {i}: таймер сдачи должен быть от 1 до 720 минут",
            )

        target_tag_id = int(t_target_tag_raw) if t_target_tag_raw.isdigit() else None
        if target_tag_id is not None and not db.query(Tag.id).filter(Tag.id == target_tag_id).first():
            raise HTTPException(status_code=422, detail=f"Билет {i}: выбранный тег не найден")

        if t_opens_raw or t_closes_raw:
            if not t_opens_raw:
                raise HTTPException(status_code=422, detail=f"Билет {i}: укажите время открытия")
            if not t_closes_raw:
                raise HTTPException(status_code=422, detail=f"Билет {i}: укажите время закрытия")
            opens_at = _parse_msk_datetime_local(
                t_opens_raw, ticket_number=i, field_label="открывается"
            )
            closes_at = _parse_msk_datetime_local(
                t_closes_raw, ticket_number=i, field_label="закрывается"
            )
            opens_msk = opens_at.astimezone(MSK_TZ)
            closes_msk = closes_at.astimezone(MSK_TZ)
            if closes_at <= opens_at:
                raise HTTPException(
                    status_code=422,
                    detail=f"Билет {i}: время закрытия должно быть позже открытия",
                )
            if (
                restrict_start_by_duration
                and closes_msk - timedelta(minutes=duration_minutes) < opens_msk
            ):
                raise HTTPException(
                    status_code=422,
                    detail=(
                        f"Билет {i}: период должен быть не короче времени на "
                        f"выполнение ({duration_minutes} мин) — иначе билет нельзя "
                        f"будет получить"
                    ),
                )
            if closes_at < datetime.now(timezone.utc):
                raise HTTPException(status_code=422, detail=f"Билет {i}: время закрытия уже в прошлом")
            t_start = opens_msk.date()
            t_end = closes_msk.date()
        else:
            opens_at = None
            closes_at = None
            try:
                t_end = date.fromisoformat(t_end_raw)
            except ValueError:
                raise HTTPException(status_code=422, detail=f"Неверная дата окончания в билете {i}")
            if t_activate_mode == "now":
                t_start = today_msk()
            else:
                try:
                    t_start = date.fromisoformat(t_start_raw)
                except ValueError:
                    raise HTTPException(status_code=422, detail=f"Неверная дата начала в билете {i}")
            if t_end < t_start:
                raise HTTPException(status_code=422, detail=f"Дата окончания раньше начала в билете {i}")
            if t_end < today_msk():
                raise HTTPException(status_code=422, detail=f"Билет {i}: дата окончания уже в прошлом")

        student_ids = list({int(x) for x in t_students_raw.split(",") if x.strip().isdigit()})
        # Тег необязателен: если тег не выбран и не указаны конкретные ученики —
        # билет выдаётся всем (assign_to_all). Тег только сужает выдачу.
        assign_all = target_tag_id is None and (t_all or not student_ids)

        ticket = ExamTicket(
            assignment_id=assignment.id,
            ticket_number=i,
            title=t_title,
            description=t_desc,
            image_s3_url=t_img_url,
            image_s3_path=t_img_path,
            start_date=t_start,
            end_date=t_end,
            opens_at=opens_at,
            closes_at=closes_at,
            duration_minutes=duration_minutes,
            restrict_start_by_duration=restrict_start_by_duration,
            target_tag_id=target_tag_id,
            assign_to_all=assign_all,
        )
        db.add(ticket)
        db.flush()

        if target_tag_id is None and not t_all:
            if student_ids:
                stmt = pg_insert(ExamTicketAssignee.__table__).values(
                    [{"ticket_id": ticket.id, "user_id": uid} for uid in student_ids]
                ).on_conflict_do_nothing(index_elements=["ticket_id", "user_id"])
                db.execute(stmt)


@router.post("/exam-assignments/{assignment_id}/edit")
async def exam_assignment_edit_submit(
    assignment_id: int,
    request: Request,
    user: Annotated[dict, Depends(require_admin_role)],
    db: Annotated[DBSession, Depends(get_db)],
    _csrf: Annotated[None, Depends(require_csrf)],
):
    assignment = db.query(ExamAssignment).filter(ExamAssignment.id == assignment_id).first()
    if not assignment:
        raise HTTPException(status_code=404, detail="Задание не найдено")

    form = await request.form()
    kind = str(form.get("kind", "mock")).strip()
    subject = str(form.get("subject", "")).strip()
    note = str(form.get("note", "")).strip() or None
    ticket_count = int(form.get("ticket_count", 1) or 1)
    ticket_count = max(1, min(10, ticket_count))

    if kind not in ASSIGNMENT_KINDS:
        raise HTTPException(status_code=422, detail="Неверный тип задания")
    if subject not in MOCK_SUBJECTS:
        raise HTTPException(status_code=422, detail="Неверный предмет")

    # Обновляем метаданные задания. seq_number НЕ пересчитываем (монотонный, присвоен
    # при создании); название пересобираем с исходной датой создания.
    created_date = (
        assignment.created_at.astimezone(MSK_TZ).date()
        if assignment.created_at else today_msk()
    )
    assignment.kind = kind
    assignment.subject = subject
    assignment.note = note
    assignment.title = _compose_assignment_title(
        kind, assignment.seq_number, subject, created_date, note
    )

    # Full-replace тикетов: удаляем старые assignees + tickets, затем создаём из формы.
    # Notification history (notified_at) теряется — это допустимо при редактировании.
    # ExamCycle.ticket_id обнуляем вручную — FK без ON DELETE SET NULL, иначе IntegrityError.
    old_ticket_ids = [tid for (tid,) in db.query(ExamTicket.id).filter(ExamTicket.assignment_id == assignment_id).all()]
    if old_ticket_ids:
        db.query(ExamCycle).filter(ExamCycle.ticket_id.in_(old_ticket_ids)).update(
            {"ticket_id": None}, synchronize_session=False
        )
        db.query(ExamTicketAssignee).filter(ExamTicketAssignee.ticket_id.in_(old_ticket_ids)).delete(synchronize_session=False)
        db.query(ExamTicket).filter(ExamTicket.id.in_(old_ticket_ids)).delete(synchronize_session=False)
        db.flush()

    _build_tickets_from_form(db, assignment, form, ticket_count)
    _ensure_mock_exam_period_open(db, user["user_id"])
    db.commit()
    return RedirectResponse(
        f"/cabinet/exam-assignments/{assignment_id}", status_code=303
    )


# ── Просмотр задания ─────────────────────────────────────────────────────────

@router.get("/exam-assignments/{assignment_id}", response_class=HTMLResponse)
def exam_assignment_detail(
    assignment_id: int,
    request: Request,
    user: Annotated[dict, Depends(require_admin_role)],
    db: Annotated[DBSession, Depends(get_db)],
):
    assignment = db.query(ExamAssignment).filter(ExamAssignment.id == assignment_id).first()
    if not assignment:
        raise HTTPException(status_code=404, detail="Задание не найдено")

    tickets = (
        db.query(ExamTicket)
        .filter(ExamTicket.assignment_id == assignment_id)
        .order_by(ExamTicket.ticket_number)
        .all()
    )

    # Load all assignees in one query, then group in memory (avoids N+1)
    assignees_by_ticket: dict[int, list[dict]] = {t.id: [] for t in tickets}
    non_all_ticket_ids = [t.id for t in tickets if not t.assign_to_all]
    if non_all_ticket_ids:
        rows = (
            db.query(User.id, User.first_name, User.last_name, User.name,
                     ExamTicketAssignee.ticket_id, ExamTicketAssignee.notified_at)
            .join(ExamTicketAssignee, ExamTicketAssignee.user_id == User.id)
            .filter(ExamTicketAssignee.ticket_id.in_(non_all_ticket_ids))
            .all()
        )
        for r in rows:
            assignees_by_ticket.setdefault(r.ticket_id, []).append({
                "id": r.id,
                "name": f"{r.last_name or ''} {r.first_name or r.name}".strip(),
                "notified": r.notified_at is not None,
            })

    creator = db.query(User).filter(User.id == assignment.created_by_id).first()
    target_tag_ids = [t.target_tag_id for t in tickets if t.target_tag_id is not None]
    tags_by_id = {
        tag.id: tag.name
        for tag in db.query(Tag).filter(Tag.id.in_(target_tag_ids)).all()
    } if target_tag_ids else {}
    ticket_meta = {
        t.id: {
            "opens_at": _format_msk_display(ticket_opens_at(t)),
            "closes_at": _format_msk_display(ticket_closes_at(t)),
            "latest_start_at": _format_msk_display(ticket_start_cutoff_at(t)),
            "duration_minutes": ticket_duration_sec(t) // 60,
            "target_tag_name": tags_by_id.get(t.target_tag_id),
        }
        for t in tickets
    }

    return templates.TemplateResponse("superadmin_exam_assignment_detail.html", {
        "request": request,
        "user": user,
        "assignment": assignment,
        "tickets": tickets,
        "ticket_meta": ticket_meta,
        "assignees_by_ticket": assignees_by_ticket,
        "creator": creator,
    })


# ── Архивирование задания ────────────────────────────────────────────────────

@router.post("/exam-assignments/{assignment_id}/archive")
def exam_assignment_archive(
    assignment_id: int,
    user: Annotated[dict, Depends(require_admin_role)],
    db: Annotated[DBSession, Depends(get_db)],
    _csrf: Annotated[None, Depends(require_csrf)],
):
    assignment = db.query(ExamAssignment).filter(ExamAssignment.id == assignment_id).first()
    if not assignment:
        raise HTTPException(status_code=404, detail="Задание не найдено")
    assignment.status = "archived"
    db.commit()
    return RedirectResponse("/cabinet/exam-assignments/active", status_code=303)


# ── Вкл/Выкл и возврат из архива ─────────────────────────────────────────────

@router.post("/exam-assignments/{assignment_id}/toggle")
def exam_assignment_toggle(
    assignment_id: int,
    user: Annotated[dict, Depends(require_admin_role)],
    db: Annotated[DBSession, Depends(get_db)],
    _csrf: Annotated[None, Depends(require_csrf)],
):
    """Переключает пробник Вкл↔Выкл. published = принимаем сдачу, draft = пауза
    (get_active_ticket вернёт None → сдача отклоняется). Архивные не трогаем."""
    assignment = db.query(ExamAssignment).filter(ExamAssignment.id == assignment_id).first()
    if not assignment:
        raise HTTPException(status_code=404, detail="Задание не найдено")
    if assignment.status == "published":
        assignment.status = "draft"
    elif assignment.status == "draft":
        assignment.status = "published"
    db.commit()
    return RedirectResponse("/cabinet/exam-assignments/active", status_code=303)


@router.post("/exam-assignments/{assignment_id}/activate")
def exam_assignment_activate(
    assignment_id: int,
    user: Annotated[dict, Depends(require_admin_role)],
    db: Annotated[DBSession, Depends(get_db)],
    _csrf: Annotated[None, Depends(require_csrf)],
):
    """Возвращает пробник из архива в активные (включённым)."""
    assignment = db.query(ExamAssignment).filter(ExamAssignment.id == assignment_id).first()
    if not assignment:
        raise HTTPException(status_code=404, detail="Задание не найдено")
    assignment.status = "published"
    db.commit()
    return RedirectResponse("/cabinet/exam-assignments/active", status_code=303)


# ── Дублирование задания ─────────────────────────────────────────────────────

@router.post("/exam-assignments/{assignment_id}/duplicate")
def exam_assignment_duplicate(
    assignment_id: int,
    user: Annotated[dict, Depends(require_admin_role)],
    db: Annotated[DBSession, Depends(get_db)],
    _csrf: Annotated[None, Depends(require_csrf)],
):
    """Создаёт копию задания (как черновик) со всеми билетами и назначениями,
    чтобы переиспользовать пробник под новые даты/учеников.

    Копия всегда draft (Выключен), чтобы не уйти в эфир со старыми датами.
    Редиректим на форму редактирования копии — там главный преподаватель правит настройки и
    включает вручную. Работает и для активных, и для архивных заданий.
    """
    source = db.query(ExamAssignment).filter(ExamAssignment.id == assignment_id).first()
    if not source:
        raise HTTPException(status_code=404, detail="Задание не найдено")

    # " (копия)" с защитой от переполнения String(200)
    suffix = " (копия)"
    copy = ExamAssignment(
        title=source.title[: 200 - len(suffix)] + suffix,
        subject=source.subject,
        created_by_id=user["user_id"],
        status="draft",
    )
    db.add(copy)
    db.flush()

    src_tickets = (
        db.query(ExamTicket)
        .filter(ExamTicket.assignment_id == assignment_id)
        .order_by(ExamTicket.ticket_number)
        .all()
    )

    # Назначения грузим одним запросом и группируем (без N+1)
    assignees_by_ticket: dict[int, list[int]] = {}
    non_all_ids = [t.id for t in src_tickets if not t.assign_to_all]
    if non_all_ids:
        for tid, uid in (
            db.query(ExamTicketAssignee.ticket_id, ExamTicketAssignee.user_id)
            .filter(ExamTicketAssignee.ticket_id.in_(non_all_ids))
            .all()
        ):
            assignees_by_ticket.setdefault(tid, []).append(uid)

    for st in src_tickets:
        new_ticket = ExamTicket(
            assignment_id=copy.id,
            ticket_number=st.ticket_number,
            title=st.title,
            description=st.description,
            image_s3_url=st.image_s3_url,
            image_s3_path=st.image_s3_path,
            start_date=st.start_date,
            end_date=st.end_date,
            opens_at=st.opens_at,
            closes_at=st.closes_at,
            duration_minutes=st.duration_minutes,
            target_tag_id=st.target_tag_id,
            assign_to_all=st.assign_to_all,
        )
        db.add(new_ticket)
        db.flush()
        # notified_at не копируем — у копии ещё никто не уведомлён
        db.add_all([
            ExamTicketAssignee(ticket_id=new_ticket.id, user_id=uid)
            for uid in assignees_by_ticket.get(st.id, [])
        ])

    db.commit()
    return RedirectResponse(
        f"/cabinet/exam-assignments/{copy.id}/edit", status_code=303
    )


# ── Feature periods ───────────────────────────────────────────────────────────

ALL_FEATURES = [FEATURE_PORTFOLIO_UPLOAD, FEATURE_MOCK_EXAM, FEATURE_RETAKE]


@router.get("/periods", response_class=HTMLResponse)
def periods_list(
    request: Request,
    user: Annotated[dict, Depends(require_admin_role)],
    db: Annotated[DBSession, Depends(get_db)],
):
    periods = (
        db.query(FeaturePeriod)
        .order_by(FeaturePeriod.feature, FeaturePeriod.start_date.desc())
        .all()
    )
    today = today_msk()
    return templates.TemplateResponse("periods_management.html", {
        "request": request,
        "user": user,
        "periods": periods,
        "features": ALL_FEATURES,
        "feature_labels": FEATURE_LABELS,
        "today": today,
    })


@router.post("/periods/create")
def period_create(
    request: Request,
    user: Annotated[dict, Depends(require_admin_role)],
    db: Annotated[DBSession, Depends(get_db)],
    _csrf: Annotated[None, Depends(require_csrf)],
    feature: Annotated[str, Form()],
    start_date: Annotated[str, Form()],
    end_date: Annotated[str, Form()],
    title: Annotated[str, Form()] = "",
):
    if feature not in ALL_FEATURES:
        raise HTTPException(status_code=400, detail="Неверный тип периода")
    try:
        from datetime import date as _date
        sd = _date.fromisoformat(start_date)
        ed = _date.fromisoformat(end_date)
    except ValueError:
        raise HTTPException(status_code=400, detail="Неверный формат даты")
    if sd > ed:
        raise HTTPException(status_code=400, detail="Дата начала должна быть раньше даты окончания")

    period = FeaturePeriod(
        feature=feature,
        title=title.strip() or None,
        start_date=sd,
        end_date=ed,
        created_by_id=user["user_id"],
    )
    db.add(period)
    db.commit()
    invalidate_feature_cache(feature)
    return RedirectResponse("/cabinet/periods", status_code=303)


@router.post("/periods/{period_id}/deactivate")
def period_deactivate(
    period_id: int,
    user: Annotated[dict, Depends(require_admin_role)],
    db: Annotated[DBSession, Depends(get_db)],
    _csrf: Annotated[None, Depends(require_csrf)],
):
    period = db.query(FeaturePeriod).filter(FeaturePeriod.id == period_id).first()
    if not period:
        raise HTTPException(status_code=404, detail="Период не найден")
    period.is_active = False
    db.commit()
    invalidate_feature_cache(period.feature)
    return RedirectResponse("/cabinet/periods", status_code=303)


@router.post("/periods/{period_id}/activate")
def period_activate(
    period_id: int,
    user: Annotated[dict, Depends(require_admin_role)],
    db: Annotated[DBSession, Depends(get_db)],
    _csrf: Annotated[None, Depends(require_csrf)],
):
    period = db.query(FeaturePeriod).filter(FeaturePeriod.id == period_id).first()
    if not period:
        raise HTTPException(status_code=404, detail="Период не найден")
    period.is_active = True
    db.commit()
    invalidate_feature_cache(period.feature)
    return RedirectResponse("/cabinet/periods", status_code=303)


@router.post("/periods/quick-toggle")
def period_quick_toggle(
    user: Annotated[dict, Depends(require_admin_role)],
    db: Annotated[DBSession, Depends(get_db)],
    _csrf: Annotated[None, Depends(require_csrf)],
    feature: Annotated[str, Form()],
    redirect_to: Annotated[str, Form()] = "/cabinet",
):
    """Быстрое включение/выключение периода с дашборда.

    Если есть активный период на сегодня → деактивирует его.
    Если нет → создаёт период «сейчас + 30 дней».
    """
    if feature not in ALL_FEATURES:
        raise HTTPException(status_code=400, detail="Неверный тип периода")

    from datetime import timedelta
    today = today_msk()
    active_period = (
        db.query(FeaturePeriod)
        .filter(
            FeaturePeriod.feature == feature,
            FeaturePeriod.is_active.is_(True),
            FeaturePeriod.start_date <= today,
            FeaturePeriod.end_date >= today,
        )
        .first()
    )

    if active_period:
        active_period.is_active = False
    else:
        db.add(FeaturePeriod(
            feature=feature,
            title="Быстрый доступ",
            start_date=today,
            end_date=today + timedelta(days=30),
            created_by_id=user["user_id"],
        ))

    db.commit()
    invalidate_feature_cache(feature)

    safe_redirect = redirect_to if redirect_to.startswith("/") and not redirect_to.startswith("//") else "/cabinet"
    return RedirectResponse(safe_redirect, status_code=303)


# ═══════════════════════════════════════════════════════════════════════════════
# Статистика периодов сдачи
# ═══════════════════════════════════════════════════════════════════════════════

from app.services.period_stats import (
    get_mock_subject_status,
    get_ticket_receipt_stats,
    get_mock_feedback_rows,
    get_mock_score_stats,
    get_all_periods as _get_all_periods,
)


@router.get("/superadmin/activity", response_class=HTMLResponse)
def superadmin_activity(
    request: Request,
    user: Annotated[dict, Depends(require_admin_role)],
    db: Annotated[DBSession, Depends(get_db)],
):
    """Статистика активности: логины, скорость проверки, реакция на уведомления,
    возвраты на правку, онбординг, просмотры видео-отчётов, журнал изменений."""
    from app.services.activity_stats import (
        get_audit_feed,
        get_curator_review_speed,
        get_cycle_duration_stats,
        get_feedback_curator_stats,
        get_login_link_stats,
        get_login_stats,
        get_mock_attempt_stats,
        get_notification_reaction,
        get_onboarding_funnel,
        get_report_view_stats,
        get_retake_stats,
        get_revision_stats,
        get_self_score_stats,
    )

    return templates.TemplateResponse("superadmin_activity.html", {
        "request": request,
        "user": user,
        "logins": get_login_stats(db),
        "review_speed": get_curator_review_speed(db),
        "feedback_curators": get_feedback_curator_stats(db),
        "notifications": get_notification_reaction(db),
        "revisions": get_revision_stats(db),
        "onboarding": get_onboarding_funnel(db),
        "report_views": get_report_view_stats(db),
        "mock_attempts": get_mock_attempt_stats(db),
        "cycles": get_cycle_duration_stats(db),
        "retakes": get_retake_stats(db),
        "login_links": get_login_link_stats(db),
        "self_scores": get_self_score_stats(db),
        "audit_feed": get_audit_feed(db),
    })


@router.get("/superadmin/stats", response_class=HTMLResponse)
def superadmin_stats(
    request: Request,
    user: Annotated[dict, Depends(require_admin_role)],
    db: Annotated[DBSession, Depends(get_db)],
    period_id: int | None = None,
):
    # Страница только про пробники → периоды фильтруем до mock_exam.
    periods = [p for p in _get_all_periods(db) if p.feature == "mock_exam"]
    status = get_mock_subject_status(db, period_id=period_id)
    ticket_stats = get_ticket_receipt_stats(db, period_id=period_id)
    feedback_table = get_mock_feedback_rows(db, period_id=period_id)
    score_stats = get_mock_score_stats(db, period_id=period_id)
    return templates.TemplateResponse("superadmin_stats.html", {
        "request": request,
        "user": user,
        "periods": periods,
        "status": status,
        "ticket_stats": ticket_stats,
        "feedback_table": feedback_table,
        "score_stats": score_stats,
        "selected_period_id": period_id,
        "tariffs": TARIFFS,
    })


@router.get("/superadmin/stats/export")
def superadmin_stats_export(
    user: Annotated[dict, Depends(require_admin_role)],
    db: Annotated[DBSession, Depends(get_db)],
    period_id: int | None = None,
):
    import io
    import openpyxl
    from openpyxl.styles import Font, PatternFill, Alignment
    from fastapi.responses import StreamingResponse

    status = get_mock_subject_status(db, period_id=period_id)
    ticket_stats = get_ticket_receipt_stats(db, period_id=period_id)
    # limit=None: в Excel выгружаем всю таблицу ОС (страница капается 500-ю).
    feedback_table = get_mock_feedback_rows(db, period_id=period_id, limit=None)
    by_tariff_mock: dict = status["by_tariff_mock_status"]
    not_submitted_by_tariff: dict = status["not_submitted_by_tariff"]

    HEADER_FONT = Font(bold=True, color="FFFFFF")
    HEADER_FILL = PatternFill(fill_type="solid", fgColor="2563EB")
    FILL_YES = PatternFill(fill_type="solid", fgColor="D1FAE5")   # зелёный
    FILL_NO  = PatternFill(fill_type="solid", fgColor="FEE2E2")   # красный
    CENTER = Alignment(horizontal="center")

    def _style_header_row(ws, columns: list[str]) -> None:
        ws.append(columns)
        for cell in ws[ws.max_row]:
            cell.font = HEADER_FONT
            cell.fill = HEADER_FILL
            cell.alignment = CENTER

    def _set_col_widths(ws, widths: list[int]) -> None:
        for i, w in enumerate(widths, 1):
            ws.column_dimensions[openpyxl.utils.get_column_letter(i)].width = w

    def _fmt_tg(tg: str) -> str:
        if not tg:
            return "—"
        return tg if tg.startswith("@") else "@" + tg

    wb = openpyxl.Workbook()

    # ── Лист 1: Сводка ────────────────────────────────────────────────────────
    ws_summary = wb.active
    ws_summary.title = "Сводка"

    selected_period = (
        db.query(FeaturePeriod).filter(FeaturePeriod.id == period_id).first()
        if period_id else None
    )
    if selected_period:
        ws_summary.append([f"Период: {selected_period.title or ''}  "
                           f"{selected_period.start_date.strftime('%d.%m.%Y')} – "
                           f"{selected_period.end_date.strftime('%d.%m.%Y')}"])
    else:
        ws_summary.append(["Статистика пробников · данные с 13.06.2026"])
    ws_summary.append([])

    _style_header_row(ws_summary, ["Тариф", "Всего учеников", "Сдали оба предмета", "Не сдали хотя бы один"])
    for tariff in TARIFFS:
        all_students = by_tariff_mock.get(tariff, [])
        ns = not_submitted_by_tariff.get(tariff, [])
        submitted_both = len(all_students) - len(ns)
        ws_summary.append([tariff, len(all_students), submitted_both, len(ns)])

    # Пробники по предмету: сколько учеников сдали Рисунок / Композицию.
    risunok_cnt = sum(1 for students in by_tariff_mock.values() for s in students if s["risunok"])
    kompoziciya_cnt = sum(1 for students in by_tariff_mock.values() for s in students if s["kompoziciya"])
    if risunok_cnt or kompoziciya_cnt:
        ws_summary.append([])
        _style_header_row(ws_summary, ["Сдали по предмету", "Кол-во учеников", "", ""])
        ws_summary.append(["Рисунок", risunok_cnt, "", ""])
        ws_summary.append(["Композиция", kompoziciya_cnt, "", ""])

    _set_col_widths(ws_summary, [22, 18, 22, 26])

    # ── Листы 2-4: Все ученики по тарифам (статус пробников) ─────────────────
    for tariff in TARIFFS:
        students = by_tariff_mock.get(tariff, [])
        ws = wb.create_sheet(title=tariff[:31])
        _style_header_row(ws, ["Ученик", "VK ID", "Telegram", "Рисунок", "Композиция"])
        for s in students:
            row = [s["student_name"], s["vk_id"], _fmt_tg(s["tg_username"]),
                   "сдал" if s["risunok"] else "не сдал",
                   "сдал" if s["kompoziciya"] else "не сдал"]
            ws.append(row)
            r_idx = ws.max_row
            # Цветовая заливка ячеек Рисунок / Композиция
            ws.cell(r_idx, 4).fill = FILL_YES if s["risunok"] else FILL_NO
            ws.cell(r_idx, 5).fill = FILL_YES if s["kompoziciya"] else FILL_NO
        _set_col_widths(ws, [30, 14, 22, 12, 14])

    # ── Листы 5-7: Не сдали хотя бы один предмет ─────────────────────────────
    for tariff in TARIFFS:
        ns_students = not_submitted_by_tariff.get(tariff, [])
        sheet_name = f"Не сдали — {tariff}"[:31]
        ws_ns = wb.create_sheet(title=sheet_name)
        _style_header_row(ws_ns, ["Ученик", "VK ID", "Telegram", "Рисунок", "Композиция"])
        for s in ns_students:
            row = [s["student_name"], s["vk_id"], _fmt_tg(s["tg_username"]),
                   "сдал" if s["risunok"] else "не сдал",
                   "сдал" if s["kompoziciya"] else "не сдал"]
            ws_ns.append(row)
            r_idx = ws_ns.max_row
            ws_ns.cell(r_idx, 4).fill = FILL_YES if s["risunok"] else FILL_NO
            ws_ns.cell(r_idx, 5).fill = FILL_YES if s["kompoziciya"] else FILL_NO
        _set_col_widths(ws_ns, [30, 14, 22, 12, 14])

    # ── Листы: Полученные билеты (только для mock_exam / всех работ) ──────────
    if ticket_stats.get("applicable") and ticket_stats.get("total_receipts"):
        ws_tickets = wb.create_sheet(title="Билеты — сводка")
        _style_header_row(ws_tickets, ["Билет", "Предмет", "Учеников", "Выдач", "Последняя выдача"])
        for t in ticket_stats["by_ticket"]:
            title = t["ticket_title"] + (" (удалён)" if t["deleted"] else "")
            last_at = t["last_at"].strftime("%d.%m.%Y %H:%M") if t["last_at"] else "—"
            ws_tickets.append([title, t["subject"] or "—", t["student_count"], t["attempt_count"], last_at])
        _set_col_widths(ws_tickets, [40, 16, 12, 10, 20])

        ws_receipts = wb.create_sheet(title="Билеты — кто и когда")
        _style_header_row(ws_receipts, ["Ученик", "Билет", "Предмет", "Дата и время (МСК)"])
        for r in ticket_stats["receipts"]:
            started = r["started_at"].strftime("%d.%m.%Y %H:%M") if r["started_at"] else "—"
            ws_receipts.append([r["student_name"], r["ticket_title"], r["subject"] or "—", started])
        _set_col_widths(ws_receipts, [30, 40, 16, 20])

    # ── Лист: Пробники + обратная связь ──────────────────────────────────────
    if feedback_table.get("applicable") and feedback_table.get("total"):
        ws_fb = wb.create_sheet(title="Пробники + ОС")
        _style_header_row(ws_fb, [
            "Ученик", "Предмет", "Билет", "Балл", "Куратор",
            "Время сдачи (МСК)", "Время ОС (МСК)", "Обратная связь",
        ])
        for r in feedback_table["rows"]:
            submitted = r["submitted_at"].strftime("%d.%m.%Y %H:%M") if r["submitted_at"] else "—"
            fb_at = r["feedback_at"].strftime("%d.%m.%Y %H:%M") if r["feedback_at"] else "—"
            ws_fb.append([
                r["student_name"],
                r["subject"] or "—",
                r["ticket_title"] or "—",
                int(r["score"]) if r["score"] is not None else "—",
                r["curator_name"] or "—",
                submitted,
                fb_at,
                r["feedback_text"] or "—",
            ])
            ws_fb.cell(ws_fb.max_row, 8).alignment = Alignment(wrap_text=True, vertical="top")
        _set_col_widths(ws_fb, [28, 14, 32, 8, 26, 18, 18, 80])

    output = io.BytesIO()
    wb.save(output)
    output.seek(0)

    from urllib.parse import quote
    fname = "статистика.xlsx"
    return StreamingResponse(
        output,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": f"attachment; filename*=UTF-8''{quote(fname)}"},
    )


# ═══════════════════════════════════════════════════════════════════════════════
# Управление пользователями
# ═══════════════════════════════════════════════════════════════════════════════

from app.services.user_management import (
    can_assign_role_rank,
    can_manage_user_by_rank,
    get_curator_for_assignment,
    log_curator_change,
    log_tariff_change,
    soft_delete_user,
    toggle_user_active,
)


_SU_PAGE_SIZE = 10

# Order users in /cabinet/superadmin/users: students first, then curators, then everyone else
_SU_ROLE_GROUP_ORDER = case(
    (Role.rank == 1, 0),
    (Role.rank == 2, 1),
    else_=2,
)


def _su_role_group_rank(u: User) -> int:
    rank = u.role.rank if u.role else None
    if rank == 1:
        return 0
    if rank == 2:
        return 1
    return 2


def _wants_json_response(request: Request) -> bool:
    accept = request.headers.get("accept", "")
    requested_with = request.headers.get("x-requested-with", "").lower()
    return requested_with == "xmlhttprequest" or "application/json" in accept


def _normalize_tg_username(raw: str) -> str:
    return raw.strip().lstrip("@").lower()


def _split_tg_usernames(raw: str) -> list[str]:
    seen: set[str] = set()
    usernames: list[str] = []
    for item in raw.replace(",", "\n").replace(";", "\n").splitlines():
        username = _normalize_tg_username(item)
        if username and username not in seen:
            usernames.append(username)
            seen.add(username)
    return usernames


def _find_student_by_tg_username(db: DBSession, tg_username: str) -> User | None:
    """Find an existing student profile by Telegram username.

    tg_username is encrypted at rest, so equality filtering in SQL is not
    reliable. Load student candidates and compare normalized plaintext values
    after SQLAlchemy decrypts them.
    """
    username = _normalize_tg_username(tg_username)
    if not username:
        return None
    candidates = (
        db.query(User)
        .join(Role, User.role_id == Role.id)
        .filter(Role.rank == 1, User.deleted_at.is_(None))
        .order_by(User.profile_completed.desc(), User.updated_at.desc(), User.id.desc())
        .limit(5000)
        .all()
    )
    for candidate in candidates:
        if _normalize_tg_username(candidate.tg_username or "") == username:
            return candidate
    return None


def _load_superadmin_curators(db: DBSession) -> list[User]:
    return (
        db.query(User)
        .join(Role, User.role_id == Role.id)
        .filter(Role.rank == 2, User.is_active == True, User.deleted_at.is_(None))  # noqa: E712
        .order_by(User.last_name, User.first_name, User.name)
        .limit(500)
        .all()
    )


def _render_superadmin_create_staff(
    request: Request,
    user: dict,
    db: DBSession,
    *,
    issued_creds: dict | None = None,
    page_error: str | None = None,
):
    roles = db.query(Role).order_by(Role.rank).all()
    return templates.TemplateResponse("superadmin_create_staff.html", {
        "request": request,
        "user": user,
        "roles": roles,
        "current_user_rank": user["role_rank"],
        "issued_creds": issued_creds,
        "page_error": page_error,
    })


def _render_superadmin_assign_curator(
    request: Request,
    user: dict,
    db: DBSession,
    *,
    assignment_result: dict | None = None,
):
    curators = _load_superadmin_curators(db)
    return templates.TemplateResponse("superadmin_assign_curator.html", {
        "request": request,
        "user": user,
        "curators": curators,
        "assignment_result": assignment_result,
    })


def _render_superadmin_users(
    request: Request,
    user: dict,
    db: DBSession,
    *,
    q: str = "",
    role_rank: str = "",
    tariff: str = "",
    show_deleted: str = "",
    show_blocked: str = "",
    show_hidden: str = "",
    study_mode: str = "",
    is_publishable: str = "",
    has_case: str = "",
    curator_id: str = "",
    exam_subjects: str = "",
    tag: str = "",
    page: int = 1,
    issued_creds: dict | None = None,
    issued_link_user_id: int | None = None,
    issued_link_name: str | None = None,
    issued_link: str | None = None,
    issued_link_expires_at=None,
    issued_telegram_link_user_id: int | None = None,
    issued_telegram_link_name: str | None = None,
    issued_telegram_link: str | None = None,
    issued_telegram_link_expires_at=None,
    page_error: str | None = None,
):
    role_rank_int: int | None = int(role_rank) if role_rank.strip() else None
    show_deleted_b: bool = show_deleted.strip() in ("1", "true", "on", "yes")
    show_blocked_b: bool = show_blocked.strip() in ("1", "true", "on", "yes")
    show_hidden_b: bool = show_hidden.strip() in ("1", "true", "on", "yes")
    is_publishable_b: bool = is_publishable.strip() in ("1", "true", "on", "yes")
    has_case_b: bool = has_case.strip() in ("1", "true", "on", "yes")

    curator_id_int: int | None = None
    if curator_id.strip():
        try:
            curator_id_int = int(curator_id.strip())
        except ValueError:
            curator_id_int = None

    tag_id_int: int | None = None
    if tag.strip():
        try:
            tag_id_int = int(tag.strip())
        except ValueError:
            tag_id_int = None

    study_mode_clean = study_mode.strip().lower()
    if study_mode_clean and study_mode_clean not in STUDY_MODES:
        study_mode_clean = ""

    exam_subjects_clean = exam_subjects.strip()

    page = max(1, page)
    query = db.query(User).outerjoin(Role, User.role_id == Role.id)
    if show_blocked_b:
        query = query.filter(User.is_active == False, User.deleted_at.is_(None))  # noqa: E712
    elif not show_deleted_b:
        query = query.filter(User.deleted_at.is_(None))
    if not show_hidden_b:
        from sqlalchemy import or_ as _or
        query = query.filter(
            _or(
                Role.rank.is_(None),
                Role.rank != 1,
                User.profile_completed == True,  # noqa: E712
            )
        )
    if role_rank_int is not None:
        query = query.filter(Role.rank == role_rank_int)
    if tariff.strip():
        query = query.filter(User.tariff == tariff.strip())
    if study_mode_clean:
        query = query.filter(User.study_mode == study_mode_clean)
    if is_publishable_b:
        query = query.filter(User.is_publishable == True)  # noqa: E712
    if curator_id_int is not None:
        query = query.filter(User.curator_id == curator_id_int)
    if exam_subjects_clean:
        query = query.filter(User.exam_subjects == exam_subjects_clean)
    if tag_id_int is not None:
        query = query.join(UserTag, UserTag.user_id == User.id).filter(UserTag.tag_id == tag_id_int)

    q_clean = q.strip().lstrip("@").lower()
    if q_clean:
        like = f"%{q_clean}%"
        db_matched = query.filter(
            User.name.ilike(like) |
            User.first_name.ilike(like) |
            User.last_name.ilike(like) |
            User.staff_login.ilike(like)
        ).order_by(User.created_at.desc()).all()
        db_ids = {u.id for u in db_matched}

        all_candidates = query.order_by(User.created_at.desc()).all()
        tg_matched = [
            u for u in all_candidates
            if u.tg_username and q_clean in u.tg_username.lower()
        ]
        tg_ids = {u.id for u in tg_matched}

        merged_ids = db_ids | tg_ids
        all_users_map = {u.id: u for u in all_candidates + db_matched}
        all_filtered = [all_users_map[uid] for uid in merged_ids]
        all_filtered.sort(key=lambda u: u.created_at or u.id, reverse=True)
    elif has_case_b:
        all_filtered = query.order_by(User.created_at.desc()).all()
    else:
        all_filtered = None

    if has_case_b and all_filtered is not None:
        # in-memory case filter: загружаем пробники одним SQL-запросом
        ids = [u.id for u in all_filtered]
        case_ids: set[int] = set()
        if ids:
            rows = (
                db.query(Work.user_id, Work.subject, Work.score, Work.month,
                         Work.year, Work.scored_at, Work.created_at, Work.work_type)
                .filter(
                    Work.user_id.in_(ids),
                    Work.work_type == WORK_TYPE_MOCK_EXAM,
                    Work.status == "success",
                    Work.score.isnot(None),
                    Work.subject.isnot(None),
                )
                .all()
            )
            from collections import defaultdict as _dd
            from app.services.utils import has_case_growth as _hcg
            grouped: dict[int, list] = _dd(list)
            for r in rows:
                grouped[r.user_id].append(r)
            for uid, ws in grouped.items():
                if _hcg(ws):
                    case_ids.add(uid)
        all_filtered = [u for u in all_filtered if u.id in case_ids]

    if all_filtered is not None:
        all_filtered.sort(key=_su_role_group_rank)
        total = len(all_filtered)
        total_pages = max(1, (total + _SU_PAGE_SIZE - 1) // _SU_PAGE_SIZE)
        page = min(max(1, page), total_pages)
        users = all_filtered[(page - 1) * _SU_PAGE_SIZE: page * _SU_PAGE_SIZE]
    else:
        total = query.count()
        total_pages = max(1, (total + _SU_PAGE_SIZE - 1) // _SU_PAGE_SIZE)
        page = min(max(1, page), total_pages)
        users = (
            query.order_by(_SU_ROLE_GROUP_ORDER, User.created_at.desc())
            .offset((page - 1) * _SU_PAGE_SIZE)
            .limit(_SU_PAGE_SIZE)
            .all()
        )
    roles = db.query(Role).order_by(Role.rank).all()
    curators = _load_superadmin_curators(db)
    all_tags = get_all_tags(db)

    # has_case для отображения значка в строках (только для тех, кто на текущей странице)
    has_case_by_user: dict[int, bool] = {}
    if users:
        page_ids = [u.id for u in users]
        rows = (
            db.query(Work.user_id, Work.subject, Work.score, Work.month, Work.year,
                     Work.scored_at, Work.created_at, Work.work_type)
            .filter(
                Work.user_id.in_(page_ids),
                Work.work_type == WORK_TYPE_MOCK_EXAM,
                Work.status == "success",
                Work.score.isnot(None),
                Work.subject.isnot(None),
            )
            .all()
        )
        from collections import defaultdict as _dd
        from app.services.utils import has_case_growth as _hcg
        grouped: dict[int, list] = _dd(list)
        for r in rows:
            grouped[r.user_id].append(r)
        for uid, ws in grouped.items():
            has_case_by_user[uid] = _hcg(ws)

    return templates.TemplateResponse("superadmin_users.html", {
        "request": request,
        "user": user,
        "users": users,
        "roles": roles,
        "curators": curators,
        "tariffs": TARIFFS,
        "study_modes": STUDY_MODES,
        "study_mode_labels": STUDY_MODE_LABELS,
        "exam_subject_hints": EXAM_SUBJECT_HINTS,
        "has_case_by_user": has_case_by_user,
        "q": q,
        "role_rank": role_rank,
        "tariff": tariff,
        "show_deleted": show_deleted,
        "show_blocked": show_blocked,
        "show_hidden": "1" if show_hidden_b else "",
        "study_mode": study_mode_clean,
        "is_publishable": "1" if is_publishable_b else "",
        "has_case": "1" if has_case_b else "",
        "curator_id": curator_id,
        "exam_subjects": exam_subjects_clean,
        "tag": tag,
        "all_tags": all_tags,
        "current_user_id": user["user_id"],
        "current_user_rank": user["role_rank"],
        "page": page,
        "total_pages": total_pages,
        "total": total,
        "issued_creds": issued_creds,
        "issued_link_user_id": issued_link_user_id,
        "issued_link_name": issued_link_name,
        "issued_link": issued_link,
        "issued_link_expires_at": issued_link_expires_at,
        "issued_telegram_link_user_id": issued_telegram_link_user_id,
        "issued_telegram_link_name": issued_telegram_link_name,
        "issued_telegram_link": issued_telegram_link,
        "issued_telegram_link_expires_at": issued_telegram_link_expires_at,
        "page_error": page_error,
    })


@router.get("/superadmin/create-staff", response_class=HTMLResponse)
def superadmin_create_staff_page(
    request: Request,
    user: Annotated[dict, Depends(require_superadmin)],
    db: Annotated[DBSession, Depends(get_db)],
):
    return _render_superadmin_create_staff(request, user, db)


@router.get("/superadmin/assign-curator", response_class=HTMLResponse)
def superadmin_assign_curator_page(
    request: Request,
    user: Annotated[dict, Depends(require_admin_role)],
    db: Annotated[DBSession, Depends(get_db)],
):
    return _render_superadmin_assign_curator(request, user, db)


@router.get("/superadmin/users", response_class=HTMLResponse)
def superadmin_users(
    request: Request,
    user: Annotated[dict, Depends(require_admin_role)],
    db: Annotated[DBSession, Depends(get_db)],
    q: str = "",
    role_rank: str = Query(default=""),
    tariff: str = "",
    show_deleted: str = "",
    show_blocked: str = "",
    show_hidden: str = "",
    study_mode: str = "",
    is_publishable: str = "",
    has_case: str = "",
    curator_id: str = "",
    exam_subjects: str = "",
    tag: str = "",
    page: int = 1,
):
    return _render_superadmin_users(
        request,
        user,
        db,
        q=q,
        role_rank=role_rank,
        tariff=tariff,
        show_deleted=show_deleted,
        show_blocked=show_blocked,
        show_hidden=show_hidden,
        study_mode=study_mode,
        is_publishable=is_publishable,
        has_case=has_case,
        curator_id=curator_id,
        exam_subjects=exam_subjects,
        tag=tag,
        page=page,
    )


@router.post("/superadmin/users/create-student", response_class=HTMLResponse)
def superadmin_create_student(
    request: Request,
    user: Annotated[dict, Depends(require_superadmin)],
    db: Annotated[DBSession, Depends(get_db)],
    _csrf: Annotated[None, Depends(require_csrf)],
    first_name: str = Form(...),
    last_name: str = Form(""),
    tg_username: str = Form(...),
    tariff: str = Form("УВЕРЕННЫЙ"),
    curator_id: str = Form(""),
):
    first_name_clean = first_name.strip()
    last_name_clean = last_name.strip()
    tg_username_clean = _normalize_tg_username(tg_username)
    tariff_clean = tariff.strip().upper()
    if not first_name_clean:
        raise HTTPException(status_code=400, detail="Имя ученика обязательно")
    if not tg_username_clean:
        raise HTTPException(status_code=400, detail="Telegram username обязателен")
    if not _TG_USERNAME_RE.match(tg_username_clean):
        raise HTTPException(status_code=400, detail="Telegram username должен содержать 4–32 символа: латиница, цифры или _")
    if tariff_clean not in TARIFFS:
        raise HTTPException(status_code=400, detail="Неверный тариф")

    student_role = db.query(Role).filter(Role.rank == 1).first()
    if not student_role:
        raise HTTPException(status_code=400, detail="Роль ученика не найдена")

    curator_id_v: int | None = None
    curator_id_clean = curator_id.strip()
    if curator_id_clean:
        try:
            curator_id_v = int(curator_id_clean)
        except ValueError:
            raise HTTPException(status_code=400, detail="Неверный id куратора")
        if not get_curator_for_assignment(db, curator_id_v):
            raise HTTPException(status_code=400, detail="Куратор не найден")

    student = _find_student_by_tg_username(db, tg_username_clean)
    if student:
        if not student.first_name:
            student.first_name = first_name_clean
        if last_name_clean and not student.last_name:
            student.last_name = last_name_clean
        if not student.name:
            student.name = f"{student.first_name or first_name_clean} {student.last_name or last_name_clean}".strip()
        if tariff_clean:
            log_tariff_change(db, user["user_id"], student.id, student.tariff, tariff_clean)
            student.tariff = tariff_clean
        if curator_id_v is not None:
            log_curator_change(db, user["user_id"], student.id, student.curator_id, curator_id_v)
            student.curator_id = curator_id_v
        student.role_id = student_role.id
        student.is_active = True
    else:
        full_name = f"{first_name_clean} {last_name_clean}".strip()
        student = User(
            vk_id=next_manual_vk_id(db),
            name=full_name,
            first_name=first_name_clean,
            last_name=last_name_clean or None,
            tg_username=tg_username_clean,
            tariff=tariff_clean,
            role_id=student_role.id,
            curator_id=curator_id_v,
            is_active=True,
            is_group_member=False,
            profile_completed=False,
        )
        db.add(student)
        db.flush()
        if curator_id_v is not None:
            log_curator_change(db, user["user_id"], student.id, None, curator_id_v)
    db.flush()
    issued_creds = _issue_login_password(db, student)
    db.commit()

    return _render_superadmin_users(request, user, db, issued_creds=issued_creds)


@router.post("/superadmin/users/create-staff", response_class=HTMLResponse)
def superadmin_create_staff(
    request: Request,
    user: Annotated[dict, Depends(require_superadmin)],
    db: Annotated[DBSession, Depends(get_db)],
    _csrf: Annotated[None, Depends(require_csrf)],
    first_name: str = Form(...),
    last_name: str = Form(""),
    role_id: int = Form(...),
):
    first_name_clean = first_name.strip()
    last_name_clean = last_name.strip()
    if not first_name_clean:
        return _render_superadmin_create_staff(request, user, db, page_error="Имя сотрудника обязательно.")

    new_role = db.query(Role).filter(Role.id == role_id).first()
    if not new_role:
        return _render_superadmin_create_staff(request, user, db, page_error="Роль не найдена.")
    if new_role.rank < 2:
        return _render_superadmin_create_staff(request, user, db, page_error="Аккаунт сотрудника требует роль с рангом ≥ 2.")
    if not can_assign_role_rank(user["role_rank"], new_role.rank):
        return _render_superadmin_create_staff(request, user, db, page_error="Нельзя создать аккаунт с рангом не ниже вашего.")

    full_name = f"{first_name_clean} {last_name_clean}".strip()
    staff = User(
        vk_id=next_manual_vk_id(db),
        name=full_name,
        first_name=first_name_clean,
        last_name=last_name_clean or None,
        role_id=new_role.id,
        is_active=True,
        is_group_member=False,
        tariff="УВЕРЕННЫЙ",
    )
    db.add(staff)
    db.flush()
    issued_creds = _issue_login_password(db, staff)
    db.commit()

    return _render_superadmin_create_staff(request, user, db, issued_creds=issued_creds)


@router.post("/superadmin/users/{target_id}/set-credentials", response_class=HTMLResponse)
def superadmin_user_set_credentials(
    target_id: int,
    request: Request,
    user: Annotated[dict, Depends(require_admin_role)],
    db: Annotated[DBSession, Depends(get_db)],
    _csrf: Annotated[None, Depends(require_csrf)],
):
    target = db.query(User).filter(
        User.id == target_id,
        User.is_active == True,  # noqa: E712
        User.deleted_at.is_(None),
    ).first()
    if not target:
        raise HTTPException(status_code=404, detail="Пользователь не найден")

    if not can_manage_user_by_rank(user["user_id"], user["role_rank"], target):
        raise HTTPException(status_code=403, detail="Нельзя выдать доступ роли равной или выше своей")

    issued_creds = _issue_login_password(db, target)
    db.commit()
    return _render_superadmin_users(request, user, db, issued_creds=issued_creds)


@router.post("/superadmin/users/{target_id}/issue-link", response_class=HTMLResponse)
def superadmin_user_issue_link(
    target_id: int,
    request: Request,
    user: Annotated[dict, Depends(require_admin_role)],
    db: Annotated[DBSession, Depends(get_db)],
    _csrf: Annotated[None, Depends(require_csrf)],
):
    target = db.query(User).filter(User.id == target_id).first()
    if not target:
        return _render_superadmin_users(request, user, db, page_error="Пользователь не найден.")
    if not can_manage_user_by_rank(user["user_id"], user["role_rank"], target):
        return _render_superadmin_users(
            request,
            user,
            db,
            page_error="Нельзя выпустить ссылку для роли равной или выше своей.",
        )
    if not target.is_active:
        return _render_superadmin_users(
            request,
            user,
            db,
            page_error="Нельзя выпустить ссылку для неактивного пользователя.",
        )
    if not target.role_id and not target.is_group_member and not target.is_admin:
        return _render_superadmin_users(
            request,
            user,
            db,
            page_error="Одноразовая ссылка доступна только пользователям с назначенной ролью.",
        )

    base_url = f"https://{settings.domain}" if settings.domain else str(request.base_url).rstrip("/")
    issued_link, login_token = issue_one_time_login_link(
        db,
        user=target,
        base_url=base_url,
        issued_by=f"superadmin:{user['user_id']}",
    )
    return _render_superadmin_users(
        request,
        user,
        db,
        issued_link_user_id=target.id,
        issued_link_name=_display_user_name(target),
        issued_link=issued_link,
        issued_link_expires_at=login_token.expires_at,
    )


@router.post("/superadmin/users/{target_id}/issue-telegram-link", response_class=HTMLResponse)
def superadmin_user_issue_telegram_link(
    target_id: int,
    request: Request,
    user: Annotated[dict, Depends(require_admin_role)],
    db: Annotated[DBSession, Depends(get_db)],
    _csrf: Annotated[None, Depends(require_csrf)],
):
    """Ссылка-приглашение для действующего ученика привязать Telegram к его
    текущему аккаунту вместо создания нового при переходе с VK-входа —
    портфолио и оценки остаются на месте (см. auth.py::_handle_telegram_link_start)."""
    if not settings.telegram_bot_username:
        return _render_superadmin_users(request, user, db, page_error="Telegram-бот ещё не настроен.")

    target = db.query(User).filter(User.id == target_id).first()
    if not target:
        return _render_superadmin_users(request, user, db, page_error="Пользователь не найден.")
    if not can_manage_user_by_rank(user["user_id"], user["role_rank"], target):
        return _render_superadmin_users(
            request,
            user,
            db,
            page_error="Нельзя выпустить ссылку для роли равной или выше своей.",
        )
    if not target.is_active:
        return _render_superadmin_users(
            request,
            user,
            db,
            page_error="Нельзя выпустить ссылку для неактивного пользователя.",
        )

    raw_token, link_token = issue_telegram_link_token(
        db,
        user=target,
        issued_by=f"superadmin:{user['user_id']}",
    )
    deep_link = f"https://t.me/{settings.telegram_bot_username}?start={raw_token}"
    return _render_superadmin_users(
        request,
        user,
        db,
        issued_telegram_link_user_id=target.id,
        issued_telegram_link_name=_display_user_name(target),
        issued_telegram_link=deep_link,
        issued_telegram_link_expires_at=link_token.expires_at,
    )


@router.post("/superadmin/users/{target_id}/role")
def superadmin_user_set_role(
    target_id: int,
    user: Annotated[dict, Depends(require_admin_role)],
    db: Annotated[DBSession, Depends(get_db)],
    _csrf: Annotated[None, Depends(require_csrf)],
    role_id: str = Form(...),
):
    target = db.query(User).filter(User.id == target_id).first()
    if not target:
        raise HTTPException(status_code=404, detail="Пользователь не найден")

    acting_rank = user.get("role_rank", 0)
    if not can_manage_user_by_rank(user["user_id"], acting_rank, target):
        return RedirectResponse("/cabinet/superadmin/users", status_code=303)

    if role_id == "":
        target.role_id = None
        db.commit()
        _invalidate_user_sessions(db, target.id)
        return RedirectResponse("/cabinet/superadmin/users", status_code=303)

    try:
        role_id_int = int(role_id)
    except ValueError:
        raise HTTPException(status_code=400, detail="Неверная роль")

    new_role = db.query(Role).filter(Role.id == role_id_int).first()
    if not new_role:
        raise HTTPException(status_code=404, detail="Роль не найдена")
    if not can_assign_role_rank(acting_rank, new_role.rank):
        return RedirectResponse("/cabinet/superadmin/users", status_code=303)

    target.role_id = new_role.id
    db.commit()
    _invalidate_user_sessions(db, target.id)
    return RedirectResponse("/cabinet/superadmin/users", status_code=303)


@router.post("/superadmin/users/assign-curator-bulk", response_class=HTMLResponse)
def superadmin_assign_curator_bulk(
    request: Request,
    user: Annotated[dict, Depends(require_admin_role)],
    db: Annotated[DBSession, Depends(get_db)],
    _csrf: Annotated[None, Depends(require_csrf)],
    curator_id: int = Form(...),
    student_usernames: str = Form(""),
    cohort_tag: str = Form(""),
    exam_subjects: str = Form(""),
):
    cohort_tag = cohort_tag.strip().lower()
    if cohort_tag and cohort_tag not in COHORT_TAGS:
        raise HTTPException(status_code=400, detail="Неверная метка группы")

    exam_subjects_v = exam_subjects.strip() or None
    if exam_subjects_v and exam_subjects_v not in EXAM_SUBJECT_HINTS:
        raise HTTPException(status_code=400, detail="Неверный предмет")

    curator = get_curator_for_assignment(db, curator_id, active_only=True)
    if not curator:
        raise HTTPException(status_code=404, detail="Куратор не найден")

    curator_tag_v = f"{curator.last_name or ''} {curator.first_name or curator.name}".strip() or None

    usernames = _split_tg_usernames(student_usernames)
    result = {
        "curator_name": curator_tag_v,
        "requested": len(usernames),
        "matched": 0,
        "assigned": 0,
        "unchanged": 0,
        "not_found": [],
        "cohort_tag": cohort_tag,
    }
    if not usernames:
        result["error"] = "Добавьте хотя бы один Telegram username."
        return _render_superadmin_assign_curator(request, user, db, assignment_result=result)

    wanted = set(usernames)
    students = (
        db.query(User)
        .join(Role, User.role_id == Role.id)
        .filter(Role.rank == 1, User.is_active == True, User.deleted_at.is_(None))  # noqa: E712
        .all()
    )
    by_username = {
        _normalize_tg_username(s.tg_username or ""): s
        for s in students
        if _normalize_tg_username(s.tg_username or "") in wanted
    }

    for username in usernames:
        student = by_username.get(username)
        if not student:
            result["not_found"].append(username)
            continue
        result["matched"] += 1
        if student.curator_id == curator.id:
            result["unchanged"] += 1
        else:
            log_curator_change(db, user["user_id"], student.id, student.curator_id, curator.id)
            student.curator_id = curator.id
            result["assigned"] += 1
        if cohort_tag:
            student.cohort_tag = cohort_tag
        if curator_tag_v:
            student.curator_tag = curator_tag_v
        if exam_subjects_v:
            student.exam_subjects = exam_subjects_v

    if result["assigned"] or (cohort_tag and result["matched"]) \
            or (curator_tag_v and result["matched"]) \
            or (exam_subjects_v and result["matched"]):
        db.commit()
    else:
        db.rollback()

    return _render_superadmin_assign_curator(request, user, db, assignment_result=result)


def _invalidate_user_sessions(db: DBSession, user_id: int) -> None:
    """Drop Redis-cached session dicts for every active session of a user."""
    try:
        from app.models.session import Session as _SessionModel
        sessions = db.query(_SessionModel.id).filter(
            _SessionModel.user_id == user_id,
            _SessionModel.is_active == True,  # noqa: E712
        ).all()
        for (sid,) in sessions:
            _invalidate_session_cache(sid)
    except Exception as exc:  # noqa: BLE001
        logger.warning("invalidate_user_sessions failed for user_id=%s: %s", user_id, exc)


@router.get("/superadmin/users/{target_id}", response_class=HTMLResponse)
def superadmin_user_card(
    target_id: int,
    request: Request,
    user: Annotated[dict, Depends(require_admin_role)],
    db: Annotated[DBSession, Depends(get_db)],
):
    target = (
        db.query(User)
        .outerjoin(Role, User.role_id == Role.id)
        .filter(User.id == target_id)
        .first()
    )
    if not target:
        raise HTTPException(status_code=404, detail="Пользователь не найден")

    curator = None
    if target.curator_id:
        curator = db.query(User).filter(User.id == target.curator_id).first()

    curators = _load_superadmin_curators(db)
    roles = db.query(Role).order_by(Role.rank).all()

    return templates.TemplateResponse("superadmin_user_card.html", {
        "request": request,
        "user": user,
        "target": target,
        "target_curator": curator,
        "curators": curators,
        "roles": roles,
        "tariffs": TARIFFS,
        "cohort_tags": sorted(COHORT_TAGS),
        "study_modes": STUDY_MODES,
        "study_mode_labels": STUDY_MODE_LABELS,
        "exam_subject_hints": EXAM_SUBJECT_HINTS,
    })


@router.post("/superadmin/users/{target_id}/tags")
def superadmin_user_save_tags(
    target_id: int,
    user: Annotated[dict, Depends(require_admin_role)],
    db: Annotated[DBSession, Depends(get_db)],
    _csrf: Annotated[None, Depends(require_csrf)],
    exam_dates: str = Form(""),
    exam_subjects: str = Form(""),
    study_mode: str = Form(""),
    is_publishable: str = Form(""),
    curator_id: str = Form(""),
    curator_tag: str = Form(""),
    tariff: str = Form(""),
    about: str = Form(""),
    cohort_tag: str = Form(""),
):
    target = db.query(User).filter(User.id == target_id).first()
    if not target:
        raise HTTPException(status_code=404, detail="Пользователь не найден")

    exam_dates_v = exam_dates.strip()[:30] or None
    exam_subjects_v = exam_subjects.strip()[:20] or None
    study_mode_v = study_mode.strip().lower() or None
    if study_mode_v and study_mode_v not in STUDY_MODES:
        raise HTTPException(status_code=400, detail="Неверный режим обучения")
    is_publishable_v = is_publishable.strip() in ("1", "true", "on", "yes")

    curator_id_v: int | None
    curator_id_clean = curator_id.strip()
    if curator_id_clean:
        try:
            curator_id_v = int(curator_id_clean)
        except ValueError:
            raise HTTPException(status_code=400, detail="Неверный id куратора")
        if not get_curator_for_assignment(db, curator_id_v):
            raise HTTPException(status_code=400, detail="Куратор не найден")
    else:
        curator_id_v = None

    curator_tag_v = curator_tag.strip()[:100] or None

    tariff_v = tariff.strip()
    if tariff_v and tariff_v not in TARIFFS:
        raise HTTPException(status_code=400, detail="Неверный тариф")

    about_v = about.strip()[:500] or None

    cohort_tag_v = cohort_tag.strip().lower()
    if cohort_tag_v and cohort_tag_v not in COHORT_TAGS:
        raise HTTPException(status_code=400, detail="Неверная метка группы")
    cohort_tag_v = cohort_tag_v or None

    log_curator_change(db, user["user_id"], target.id, target.curator_id, curator_id_v)
    if tariff_v:
        log_tariff_change(db, user["user_id"], target.id, target.tariff, tariff_v)

    target.exam_dates = exam_dates_v
    target.exam_subjects = exam_subjects_v
    target.study_mode = study_mode_v
    target.is_publishable = is_publishable_v
    target.curator_id = curator_id_v
    target.curator_tag = curator_tag_v
    if tariff_v:
        target.tariff = tariff_v
    target.about = about_v
    target.cohort_tag = cohort_tag_v
    db.commit()

    _invalidate_user_sessions(db, target.id)

    return RedirectResponse(
        f"/cabinet/superadmin/users/{target.id}?saved=1",
        status_code=303,
    )


@router.post("/superadmin/users/{target_id}/curator")
def superadmin_user_set_curator(
    target_id: int,
    user: Annotated[dict, Depends(require_admin_role)],
    db: Annotated[DBSession, Depends(get_db)],
    _csrf: Annotated[None, Depends(require_csrf)],
    curator_id: str = Form(""),
):
    """Quick endpoint for changing curator from the users list (JSON response)."""
    target = db.query(User).filter(User.id == target_id).first()
    if not target:
        raise HTTPException(status_code=404, detail="Пользователь не найден")

    curator_id_clean = curator_id.strip()
    new_curator_id: int | None
    new_curator_name: str | None = None
    if curator_id_clean:
        try:
            new_curator_id = int(curator_id_clean)
        except ValueError:
            raise HTTPException(status_code=400, detail="Неверный id куратора")
        curator_obj = get_curator_for_assignment(db, new_curator_id)
        if not curator_obj:
            raise HTTPException(status_code=400, detail="Куратор не найден")
        new_curator_name = f"{curator_obj.last_name or ''} {curator_obj.first_name or curator_obj.name}".strip()
    else:
        new_curator_id = None

    log_curator_change(db, user["user_id"], target.id, target.curator_id, new_curator_id)
    target.curator_id = new_curator_id
    db.commit()
    _invalidate_user_sessions(db, target.id)

    return JSONResponse({
        "ok": True,
        "user_id": target.id,
        "curator_id": new_curator_id,
        "curator_name": new_curator_name,
    })


@router.post("/superadmin/users/{target_id}/cohort-tag")
def superadmin_user_set_cohort_tag(
    target_id: int,
    user: Annotated[dict, Depends(require_admin_role)],
    db: Annotated[DBSession, Depends(get_db)],
    _csrf: Annotated[None, Depends(require_csrf)],
    cohort_tag: str = Form(""),
):
    """Quick endpoint for changing the avatar banner (cohort_tag) from the users list (JSON response)."""
    target = db.query(User).filter(User.id == target_id).first()
    if not target:
        raise HTTPException(status_code=404, detail="Пользователь не найден")

    cohort_tag_v = cohort_tag.strip().lower()
    if cohort_tag_v and cohort_tag_v not in COHORT_TAGS:
        raise HTTPException(status_code=400, detail="Неверная метка группы")

    target.cohort_tag = cohort_tag_v or None
    db.commit()
    _invalidate_user_sessions(db, target.id)

    return JSONResponse({
        "ok": True,
        "user_id": target.id,
        "cohort_tag": target.cohort_tag,
    })


@router.post("/superadmin/users/{target_id}/delete")
def superadmin_delete_user(
    target_id: int,
    request: Request,
    user: Annotated[dict, Depends(require_admin_role)],
    db: Annotated[DBSession, Depends(get_db)],
    _csrf: Annotated[None, Depends(require_csrf)],
):
    ok = soft_delete_user(db, target_user_id=target_id, performed_by_id=user["user_id"])
    if not ok:
        raise HTTPException(status_code=400, detail="Невозможно удалить пользователя")
    if _wants_json_response(request):
        return JSONResponse({"ok": True, "user_id": target_id, "deleted": True})
    return RedirectResponse("/cabinet/superadmin/users", status_code=303)


@router.post("/superadmin/users/{target_id}/toggle-active")
def superadmin_toggle_active(
    target_id: int,
    request: Request,
    user: Annotated[dict, Depends(require_admin_role)],
    db: Annotated[DBSession, Depends(get_db)],
    _csrf: Annotated[None, Depends(require_csrf)],
):
    result = toggle_user_active(db, target_user_id=target_id, performed_by_id=user["user_id"])
    if result is None:
        raise HTTPException(status_code=400, detail="Невозможно изменить статус пользователя")
    if _wants_json_response(request):
        return JSONResponse({"ok": True, "user_id": target_id, "is_active": result})
    return RedirectResponse("/cabinet/superadmin/users", status_code=303)


# ── Drive-sync status & retry ────────────────────────────────────────────────

@router.get("/superadmin/drive-sync-status")
def drive_sync_status(
    user: Annotated[dict, Depends(require_admin_role)],
    db: Annotated[DBSession, Depends(get_db)],
):
    """Return counts of works by drive_status (admin+)."""
    rows = (
        db.query(Work.drive_status, func.count(Work.id))
        .group_by(Work.drive_status)
        .all()
    )
    counts = {status: cnt for status, cnt in rows}
    return JSONResponse({
        "enabled": settings.n8n_enabled,
        "pending": counts.get("pending", 0),
        "synced": counts.get("synced", 0),
        "failed": counts.get("failed", 0),
        "s3_only": counts.get("s3_only", 0),
    })


@router.post("/superadmin/works/{work_id}/retry-drive-sync")
async def retry_drive_sync(
    work_id: int,
    background_tasks: BackgroundTasks,
    user: Annotated[dict, Depends(require_admin_role)],
    db: Annotated[DBSession, Depends(get_db)],
    _csrf: Annotated[None, Depends(require_csrf)],
):
    """Re-queue a single work for Drive upload (admin+). Resets drive_status to pending."""
    if not settings.n8n_enabled:
        raise HTTPException(status_code=503, detail="Интеграция n8n отключена")

    work = db.query(Work).filter(Work.id == work_id).first()
    if not work:
        raise HTTPException(status_code=404, detail="Работа не найдена")

    if not work.s3_url and not work.s3_path:
        raise HTTPException(status_code=400, detail="У работы нет файла в хранилище — повторная отправка невозможна")

    student = db.query(User).filter(User.id == work.user_id).first()
    if not student:
        raise HTTPException(status_code=404, detail="Ученик не найден")

    # Reset status so background task will update it
    work.drive_status = "pending"
    db.commit()

    # Read bytes from S3 for re-sending to n8n
    import asyncio, httpx
    s3_url = work.s3_url
    photo_bytes: bytes | None = None
    if s3_url:
        try:
            async with httpx.AsyncClient(timeout=30) as client:
                resp = await client.get(s3_url)
                resp.raise_for_status()
                photo_bytes = resp.content
        except Exception as exc:
            logger.warning("retry_drive_sync: could not fetch S3 for work_id=%s: %s", work_id, exc)

    if not photo_bytes:
        raise HTTPException(status_code=502, detail="Не удалось скачать файл из хранилища для повторной отправки")

    from app.api.upload import _send_to_n8n_background
    from app.constants import MONTHS
    import datetime as _dt
    month = work.month or MONTHS[_dt.datetime.now().month - 1]

    user_dict = {
        "vk_id": student.vk_id,
        "name": student.name,
        "tariff": work.tariff or student.tariff or "УВЕРЕННЫЙ",
        "user_id": student.id,
        "session_id": None,
    }

    background_tasks.add_task(
        _send_to_n8n_background,
        work_queue=[(work.id, work.filename, photo_bytes, work.s3_path)],
        user=user_dict,
        month=month,
        work_type=work.work_type,
    )
    return JSONResponse({"success": True, "message": f"Работа #{work_id} поставлена в очередь на синхронизацию"})


# ═══════════════════════════════════════════════════════════════════════════════
# Имперсонация — войти как другой пользователь (только rank=5)
# ═══════════════════════════════════════════════════════════════════════════════

from itsdangerous import URLSafeTimedSerializer as _UTS, BadData as _BadData
from app.models.audit_log import AuditLog as _AuditLog
from app.models.session import Session as _DBSession

_IMPERSONATION_TTL_MIN = 30
_IMPERSONATION_COOKIE = "impersonation_original"
_IMPERSONATION_MAX_AGE = 60 * 60 * 24  # 24h max — sanity bound for restore


def _impersonation_serializer() -> _UTS:
    return _UTS(settings.session_secret, salt="impersonation-v1")


@router.get("/superadmin/curators", response_class=HTMLResponse)
def superadmin_curators_list(
    request: Request,
    user: Annotated[dict, Depends(require_admin_role)],
    db: Annotated[DBSession, Depends(get_db)],
):
    StudentAlias = aliased(User)
    rows = (
        db.query(
            User.id, User.first_name, User.last_name, User.name, User.photo_url,
            User.staff_login, User.is_active,
            func.count(StudentAlias.id).label("student_count"),
        )
        .join(Role, User.role_id == Role.id)
        .outerjoin(StudentAlias, (StudentAlias.curator_id == User.id) & (StudentAlias.is_active == True))
        .filter(Role.rank == 2, User.deleted_at.is_(None))
        .group_by(User.id, User.first_name, User.last_name, User.name, User.photo_url,
                  User.staff_login, User.is_active)
        .order_by(func.count(StudentAlias.id).desc(), User.last_name)
        .all()
    )
    curators = [
        {
            "id": r.id,
            "name": f"{r.last_name or ''} {r.first_name or r.name}".strip(),
            "photo_url": r.photo_url,
            "staff_login": r.staff_login,
            "is_active": r.is_active,
            "student_count": r.student_count,
        }
        for r in rows
    ]
    return templates.TemplateResponse("superadmin_curators.html", {
        "request": request,
        "user": user,
        "curators": curators,
    })


@router.post("/superadmin/impersonate/stop")
def superadmin_impersonate_stop(request: Request, db: Annotated[DBSession, Depends(get_db)]):
    """Аварийный выход из имперсонации.

    НЕ требует валидного `get_current_user` и НЕ проверяет роль — работает только
    по подписанной cookie `_IMPERSONATION_COOKIE`. Это нужно, чтобы кнопка
    «Выйти обратно» работала даже если имперсонируемый пользователь оказался
    заблокирован / удалён / без прав к текущей странице.
    """
    signed = request.cookies.get(_IMPERSONATION_COOKIE, "")
    impersonation_session_id = request.cookies.get("session_id", "")

    response = RedirectResponse("/cabinet", status_code=303)
    response.delete_cookie(_IMPERSONATION_COOKIE, path="/")

    if not signed:
        # Нечего восстанавливать — просто чистим cookie и редиректим.
        if impersonation_session_id:
            response.delete_cookie("session_id", path="/")
        return response

    try:
        original_session_id = _impersonation_serializer().loads(
            signed, max_age=_IMPERSONATION_MAX_AGE
        )
    except _BadData:
        if impersonation_session_id:
            response.delete_cookie("session_id", path="/")
        return response

    # Деактивируем имперсонационную сессию и пишем аудит (best effort)
    if impersonation_session_id:
        try:
            imp_sess = db.query(_DBSession).filter(_DBSession.id == impersonation_session_id).first()
            if imp_sess:
                db.add(_AuditLog(
                    action="impersonate_stop",
                    performed_by_id=imp_sess.impersonated_by_id or 0,
                    target_user_id=imp_sess.user_id,
                    details=f"session={impersonation_session_id}",
                ))
                imp_sess.is_active = False
                db.commit()
        except Exception as exc:  # noqa: BLE001
            logger.warning("impersonate_stop: cleanup failed for %s: %s", impersonation_session_id, exc)
            db.rollback()
        _invalidate_session_cache(impersonation_session_id)

    _invalidate_session_cache(original_session_id)

    response.set_cookie(
        key="session_id",
        value=original_session_id,
        httponly=True,
        secure=True,
        samesite="lax",
        max_age=settings.session_ttl_hours * 3600,
        path="/",
    )
    return response


@router.post("/superadmin/impersonate/{target_id}")
def superadmin_impersonate_start(
    target_id: int,
    request: Request,
    user: Annotated[dict, Depends(require_admin_role)],
    db: Annotated[DBSession, Depends(get_db)],
    _csrf: Annotated[None, Depends(require_csrf)],
):
    if user.get("impersonated_by_id"):
        raise HTTPException(status_code=400, detail="Уже в режиме имперсонации")

    target = (
        db.query(User)
        .outerjoin(Role, User.role_id == Role.id)
        .filter(User.id == target_id, User.is_active == True, User.deleted_at.is_(None))  # noqa: E712
        .first()
    )
    if not target:
        raise HTTPException(status_code=404, detail="Пользователь не найден")

    if not can_manage_user_by_rank(user["user_id"], user["role_rank"], target):
        raise HTTPException(status_code=403, detail="Нельзя имперсонировать роль равную или выше своей")

    original_session_id = user["session_id"]
    new_session = _DBSession(
        user_id=target.id,
        expires_at=datetime.now(timezone.utc) + timedelta(minutes=_IMPERSONATION_TTL_MIN),
        is_active=True,
        impersonated_by_id=user["user_id"],
    )
    db.add(new_session)
    db.add(_AuditLog(
        action="impersonate_start",
        performed_by_id=user["user_id"],
        target_user_id=target.id,
        details=f"session={new_session.id}",
    ))
    db.commit()
    db.refresh(new_session)

    signed = _impersonation_serializer().dumps(original_session_id)

    response = RedirectResponse("/cabinet", status_code=303)
    response.set_cookie(
        key="session_id",
        value=new_session.id,
        httponly=True,
        secure=True,
        samesite="lax",
        max_age=_IMPERSONATION_TTL_MIN * 60,
        path="/",
    )
    response.set_cookie(
        key=_IMPERSONATION_COOKIE,
        value=signed,
        httponly=True,
        secure=True,
        samesite="lax",
        max_age=_IMPERSONATION_MAX_AGE,
        path="/",
    )
    return response
